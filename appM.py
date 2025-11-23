import streamlit as st
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime
import io
import random
import calendar 
import pandas as pd
from itertools import groupby
from operator import itemgetter

# --- CONSTANTES Y CONFIGURACIÓN ---
TEAMS = ['A', 'B', 'C']
ROLES = ["Mando", "Conductor", "Bombero"]
MESES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", 
         "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

# Plantilla por defecto
DEFAULT_ROSTER = [
    {"ID_Puesto": "Jefe A",       "Nombre": "Jefe A",       "Turno": "A", "Rol": "Mando",      "SV": False},
    {"ID_Puesto": "Subjefe A",    "Nombre": "Subjefe A",    "Turno": "A", "Rol": "Mando",      "SV": False},
    {"ID_Puesto": "Cond A",       "Nombre": "Cond A",       "Turno": "A", "Rol": "Conductor", "SV": True},
    {"ID_Puesto": "Bombero A1",   "Nombre": "Bombero A1",   "Turno": "A", "Rol": "Bombero",    "SV": True},
    {"ID_Puesto": "Bombero A2",   "Nombre": "Bombero A2",   "Turno": "A", "Rol": "Bombero",    "SV": False},
    {"ID_Puesto": "Bombero A3",   "Nombre": "Bombero A3",   "Turno": "A", "Rol": "Bombero",    "SV": False},
    
    {"ID_Puesto": "Jefe B",       "Nombre": "Jefe B",       "Turno": "B", "Rol": "Mando",      "SV": False},
    {"ID_Puesto": "Subjefe B",    "Nombre": "Subjefe B",    "Turno": "B", "Rol": "Mando",      "SV": False},
    {"ID_Puesto": "Cond B",       "Nombre": "Cond B",       "Turno": "B", "Rol": "Conductor", "SV": True},
    {"ID_Puesto": "Bombero B1",   "Nombre": "Bombero B1",   "Turno": "B", "Rol": "Bombero",    "SV": True},
    {"ID_Puesto": "Bombero B2",   "Nombre": "Bombero B2",   "Turno": "B", "Rol": "Bombero",    "SV": False},
    {"ID_Puesto": "Bombero B3",   "Nombre": "Bombero B3",   "Turno": "B", "Rol": "Bombero",    "SV": False},

    {"ID_Puesto": "Jefe C",       "Nombre": "Jefe C",       "Turno": "C", "Rol": "Mando",      "SV": False},
    {"ID_Puesto": "Subjefe C",    "Nombre": "Subjefe C",    "Turno": "C", "Rol": "Mando",      "SV": False},
    {"ID_Puesto": "Cond C",       "Nombre": "Cond C",       "Turno": "C", "Rol": "Conductor", "SV": True},
    {"ID_Puesto": "Bombero C1",   "Nombre": "Bombero C1",   "Turno": "C", "Rol": "Bombero",    "SV": True},
    {"ID_Puesto": "Bombero C2",   "Nombre": "Bombero C2",   "Turno": "C", "Rol": "Bombero",    "SV": False},
    {"ID_Puesto": "Bombero C3",   "Nombre": "Bombero C3",   "Turno": "C", "Rol": "Bombero",    "SV": False},
]

# -------------------------------------------------------------------
# 1. MOTOR LÓGICO
# -------------------------------------------------------------------

def generate_base_schedule(year):
    is_leap = calendar.isleap(year)
    total_days = 366 if is_leap else 365
    status = {'A': 0, 'B': 2, 'C': 1} 
    schedule = {team: [] for team in TEAMS}
    for _ in range(total_days):
        for t in TEAMS:
            if status[t] == 0: schedule[t].append('T')
            else: schedule[t].append('L')
            status[t] = (status[t] + 1) % 3
    return schedule, total_days

def get_candidates(person_missing, roster_df, day_idx, current_schedule):
    candidates = []
    missing_role = person_missing['Rol']
    missing_turn = person_missing['Turno']
    
    for _, candidate in roster_df.iterrows():
        if candidate['Turno'] == missing_turn: continue
        cand_status = current_schedule[candidate['Nombre']][day_idx]
        if cand_status != 'L': continue 
        
        is_compatible = False
        if missing_role == "Mando":
            if candidate['Rol'] == "Mando": is_compatible = True
        elif missing_role == "Conductor":
            if candidate['Rol'] == "Conductor": is_compatible = True
            if candidate['SV']: is_compatible = True
        elif missing_role == "Bombero":
            if candidate['Rol'] == "Bombero": is_compatible = True
            if candidate['SV']: is_compatible = True
            
        if is_compatible:
            candidates.append(candidate['Nombre'])
    return candidates

def is_night_restricted(date_obj, night_periods):
    for start, end in night_periods:
        if date_obj == start or date_obj == end: return True
    return False

def is_in_night_period(day_idx, year, night_periods):
    current_date = datetime.date(year, 1, 1) + datetime.timedelta(days=day_idx)
    for start, end in night_periods:
        if start <= current_date <= end: return True
    return False

def calculate_spent_credits(roster_df, requests, year):
    base_sch, _ = generate_base_schedule(year)
    credits = {name: 0 for name in roster_df['Nombre']}
    for req in requests:
        name = req['Nombre']
        if name not in credits: continue 
        row = roster_df[roster_df['Nombre'] == name]
        if row.empty: continue
        turn = row.iloc[0]['Turno']
        start_idx = req['Inicio'].timetuple().tm_yday - 1
        end_idx = req['Fin'].timetuple().tm_yday - 1
        cost = 0
        for d in range(start_idx, end_idx + 1):
            if base_sch[turn][d] == 'T': cost += 1
        credits[name] += cost
    return credits

def get_clustered_dates(available_idxs, needed_count):
    if not available_idxs: return []
    groups = []
    for k, g in groupby(enumerate(available_idxs), lambda ix: ix[0] - ix[1]):
        groups.append(list(map(itemgetter(1), g)))
    groups.sort(key=len, reverse=True)
    selected = []
    for group in groups:
        if len(selected) < needed_count:
            take = min(len(group), needed_count - len(selected))
            selected.extend(group[:take])
        else: break
    return sorted(selected)

def validate_and_generate(roster_df, requests, year, night_periods):
    base_schedule_turn, total_days = generate_base_schedule(year)
    final_schedule = {} 
    turn_coverage_counters = {'A': 0, 'B': 0, 'C': 0}
    person_coverage_counters = {name: 0 for name in roster_df['Nombre']}
    name_to_turn = {row['Nombre']: row['Turno'] for _, row in roster_df.iterrows()}
    
    for _, row in roster_df.iterrows():
        final_schedule[row['Nombre']] = base_schedule_turn[row['Turno']].copy()

    day_vacations = {i: [] for i in range(total_days)}
    daily_error_codes = {} 
    
    natural_days_count = {name: 0 for name in roster_df['Nombre']}

    for req in requests:
        name = req['Nombre']
        start_idx = req['Inicio'].timetuple().tm_yday - 1
        end_idx = req['Fin'].timetuple().tm_yday - 1
        
        duration = (end_idx - start_idx) + 1
        natural_days_count[name] += duration
        
        for d in range(start_idx, end_idx + 1):
            if final_schedule[name][d] == 'T':
                day_vacations[d].append(name)
                final_schedule[name][d] = 'V'
            else:
                final_schedule[name][d] = 'V(L)'

    errors = []
    adjustments_log = []
    
    for name, days in natural_days_count.items():
        if days > 39:
            errors.append(f"{name}: Exceso de días naturales ({days} > 39).")

    for d in range(total_days):
        absent_people = day_vacations[d]
        if not absent_people: continue
        
        if len(absent_people) > 2:
            date_str = (datetime.date(year, 1, 1) + datetime.timedelta(days=d)).strftime("%d-%m")
            errors.append(f"{date_str}: Hay {len(absent_people)} personas de vacaciones (Máx 2).")
            daily_error_codes[d] = "RED"
            continue
            
        if len(absent_people) == 2:
            p1 = roster_df[roster_df['Nombre'] == absent_people[0]].iloc[0]
            p2 = roster_df[roster_df['Nombre'] == absent_people[1]].iloc[0]
            if p1['Turno'] == p2['Turno']:
                errors.append(f"Día {d+1}: {p1['Nombre']} y {p2['Nombre']} son del mismo turno.")
                daily_error_codes[d] = "YELLOW"

        for name_missing in absent_people:
            person_row = roster_df[roster_df['Nombre'] == name_missing].iloc[0]
            candidates = get_candidates(person_row, roster_df, d, final_schedule)
            
            if not candidates:
                errors.append(f"Día {d+1}: Sin cobertura para {name_missing}.")
                if d not in daily_error_codes: daily_error_codes[d] = "ORANGE"
                continue
                
            valid_candidates = []
            for cand in candidates:
                prev_day = final_schedule[cand][d-1] if d > 0 else 'L'
                prev_prev = final_schedule[cand][d-2] if d > 1 else 'L'
                is_prev_work = prev_day.startswith('T')
                is_prev_prev_work = prev_prev.startswith('T')
                if is_prev_work and is_prev_prev_work: continue 
                valid_candidates.append(cand)
            
            if not valid_candidates:
                date_str = (datetime.date(year, 1, 1) + datetime.timedelta(days=d)).strftime("%d-%m")
                errors.append(f"{date_str}: {name_missing} no tiene cobertura válida (Regla Máx 2T).")
                if d not in daily_error_codes: daily_error_codes[d] = "ORANGE"
                continue
            
            if d not in daily_error_codes:
                def sort_key(cand_name):
                    cand_turn = name_to_turn[cand_name]
                    return (turn_coverage_counters[cand_turn], person_coverage_counters[cand_name], random.random())
                valid_candidates.sort(key=sort_key)
                chosen = valid_candidates[0]
                chosen_turn = name_to_turn[chosen]
                final_schedule[chosen][d] = f"T*({name_missing})"
                adjustments_log.append((d, chosen, name_missing))
                turn_coverage_counters[chosen_turn] += 1
                person_coverage_counters[chosen] += 1

    fill_log = {} 
    if not errors:
        for name in roster_df['Nombre']:
            current_nat = natural_days_count[name]
            needed = 39 - current_nat
            added_dates = []
            if needed > 0:
                available_idx = [i for i, x in enumerate(final_schedule[name]) if x == 'L']
                if len(available_idx) >= needed:
                    fill_idxs = get_clustered_dates(available_idx, needed)
                    for idx in fill_idxs:
                        final_schedule[name][idx] = 'V(R)'
                        d_obj = datetime.date(year, 1, 1) + datetime.timedelta(days=idx)
                        added_dates.append(d_obj)
            fill_log[name] = added_dates

    return final_schedule, errors, person_coverage_counters, fill_log, adjustments_log, daily_error_codes

# -------------------------------------------------------------------
# 4. MOTOR AUTO-SOLVER Y REPARADOR
# -------------------------------------------------------------------
def detect_vacation_pattern(requests):
    if not requests: return 'scattered'
    durations = []
    for r in requests:
        dur = (r['Fin'] - r['Inicio']).days + 1
        durations.append(dur)
    avg_dur = sum(durations) / len(durations)
    if avg_dur >= 8: return 'block_large'
    elif avg_dur >= 4: return 'block_medium'
    else: return 'scattered'

def check_request_conflict(req, occupation_map, base_schedule_turn, roster_df, night_periods, total_days):
    start_idx = req['Inicio'].timetuple().tm_yday - 1
    end_idx = req['Fin'].timetuple().tm_yday - 1
    person = roster_df[roster_df['Nombre'] == req['Nombre']].iloc[0]
    
    if is_night_restricted(req['Inicio'], night_periods) or is_night_restricted(req['Fin'], night_periods): return "Nocturna"
    for d in range(start_idx, end_idx + 1):
        if d >= total_days: return "Fuera de rango"
        if base_schedule_turn[person['Turno']][d] == 'T':
            occupants = occupation_map[d]
            if len(occupants) >= 2: return "Max 2 Personas"
            for occ in occupants:
                if occ['Turno'] == person['Turno']: return f"Conflicto Turno con {occ['Nombre']}"
            for occ in occupants:
                if occ['Rol'] == person['Rol'] and person['Rol'] != "Bombero": return f"Conflicto Rol con {occ['Nombre']}"
    return None

def book_request(req, occupation_map, base_schedule_turn, roster_df):
    start_idx = req['Inicio'].timetuple().tm_yday - 1
    end_idx = req['Fin'].timetuple().tm_yday - 1
    person = roster_df[roster_df['Nombre'] == req['Nombre']].iloc[0].to_dict()
    for d in range(start_idx, end_idx + 1):
        if base_schedule_turn[person['Turno']][d] == 'T':
            occupation_map[d].append(person)

def force_balance_credits(final_requests, roster_df, base_schedule_turn):
    people = roster_df.to_dict('records')
    adjusted_requests = []
    person_map = {p['Nombre']: [] for p in people}
    for r in final_requests:
        if r['Nombre'] in person_map:
            person_map[r['Nombre']].append(r)
            
    for name, reqs in person_map.items():
        total_credits = 0
        person = roster_df[roster_df['Nombre'] == name].iloc[0]
        for r in reqs:
            s_idx = r['Inicio'].timetuple().tm_yday - 1
            e_idx = r['Fin'].timetuple().tm_yday - 1
            t_days_indices = []
            for d in range(s_idx, e_idx + 1):
                if base_schedule_turn[person['Turno']][d] == 'T': t_days_indices.append(d)
            cost = len(t_days_indices)
            if total_credits + cost > 13:
                allowed = 13 - total_credits
                if allowed <= 0: continue
                target_last_t = t_days_indices[allowed - 1]
                year_req = r['Fin'].year
                new_end_date = datetime.date(year_req, 1, 1) + datetime.timedelta(days=target_last_t)
                r['Fin'] = new_end_date
                total_credits += allowed
                adjusted_requests.append(r)
            else:
                total_credits += cost
                adjusted_requests.append(r)
    return adjusted_requests

# --- NUEVA FUNCIÓN V7.5 (ESTRATEGIA PIEDRAS GRANDES) ---
def run_auto_solver_fill(roster_df, year, night_periods, existing_requests):
    base_schedule_turn, total_days = generate_base_schedule(year)
    occupation_map = {i: [] for i in range(total_days)}
    
    # 1. Registrar solicitudes FIJAS (manuales o importadas)
    for req in existing_requests:
        book_request(req, occupation_map, base_schedule_turn, roster_df)
        
    final_requests = list(existing_requests)
    people = roster_df.to_dict('records')
    
    # --- ESTRATEGIA: Dividir en grupos por prioridad ---
    # 1. Mandos (Solo se cubren entre ellos -> PRIORIDAD MAXIMA)
    # 2. Conductores (Cobertura media)
    # 3. Bomberos (Cobertura total -> Llenan los huecos que sobran)
    
    group_mandos = [p for p in people if p['Rol'] == 'Mando']
    group_conductores = [p for p in people if p['Rol'] == 'Conductor']
    group_bomberos = [p for p in people if p['Rol'] == 'Bombero']
    others = [p for p in people if p['Rol'] not in ['Mando', 'Conductor', 'Bombero']] # Por si acaso
    
    # Orden de ejecución secuencial
    priority_groups = [group_mandos, group_conductores, group_bomberos, others]
    
    all_days = [datetime.date(year, 1, 1) + datetime.timedelta(days=i) for i in range(total_days)]
    
    for group in priority_groups:
        # Barajamos dentro del grupo para equidad entre pares
        random.shuffle(group)
        
        for p in group:
            # Calcular estado actual
            credits_got = 0
            natural_days_got = 0
            person_reqs = [r for r in final_requests if r['Nombre'] == p['Nombre']]
            
            for r in person_reqs:
                s_idx = r['Inicio'].timetuple().tm_yday - 1
                e_idx = r['Fin'].timetuple().tm_yday - 1
                dur = (e_idx - s_idx) + 1
                natural_days_got += dur
                for d in range(s_idx, e_idx + 1):
                    if base_schedule_turn[p['Turno']][d] == 'T': credits_got += 1
            
            if credits_got >= 13: continue
            
            # --- FASE 1: INTENTO INTELIGENTE (Patrones) ---
            pattern = detect_vacation_pattern(person_reqs)
            attempts = 0
            max_attempts = 1500
            
            while credits_got < 13 and attempts < max_attempts:
                duration = 1
                credits_needed = 13 - credits_got
                
                if credits_needed <= 2 or attempts > 500:
                    duration = 1 
                elif attempts > 200:
                    duration = 4
                else:
                    if pattern == 'block_large': duration = random.randint(7, 13)
                    elif pattern == 'block_medium': duration = random.randint(4, 7)
                    else: duration = 1
                
                if natural_days_got + duration > 39:
                    margin = 39 - natural_days_got
                    if margin <= 0: break 
                    duration = random.randint(1, margin)

                day = random.choice(all_days)
                d_idx = day.timetuple().tm_yday - 1
                is_start_T = (base_schedule_turn[p['Turno']][d_idx] == 'T')
                
                if not is_start_T and duration == 1:
                    attempts += 1; continue

                req = {"Nombre": p['Nombre'], "Inicio": day, "Fin": day + datetime.timedelta(days=duration-1)}
                
                if not check_request_conflict(req, occupation_map, base_schedule_turn, roster_df, night_periods, total_days):
                    overlap = False
                    for r in final_requests:
                        if r['Nombre'] == p['Nombre']:
                            if not (req['Inicio'] > r['Fin'] or req['Fin'] < r['Inicio']): overlap = True
                    
                    if not overlap:
                        book_request(req, occupation_map, base_schedule_turn, roster_df)
                        final_requests.append(req)
                        
                        natural_days_got += duration
                        s = req['Inicio'].timetuple().tm_yday - 1
                        e = req['Fin'].timetuple().tm_yday - 1
                        added_credits = 0
                        for d in range(s, e+1):
                            if base_schedule_turn[p['Turno']][d] == 'T': added_credits += 1
                        credits_got += added_credits
                attempts += 1
                
            # --- FASE 2: LA BARREDORA (Seguridad) ---
            if credits_got < 13:
                for d_idx in range(total_days):
                    if credits_got >= 13: break 
                    if natural_days_got >= 39: break 

                    if base_schedule_turn[p['Turno']][d_idx] != 'T': continue

                    day_obj = datetime.date(year, 1, 1) + datetime.timedelta(days=d_idx)
                    req_one_day = {"Nombre": p['Nombre'], "Inicio": day_obj, "Fin": day_obj}

                    if check_request_conflict(req_one_day, occupation_map, base_schedule_turn, roster_df, night_periods, total_days):
                        continue
                    
                    overlap = False
                    for r in final_requests:
                        if r['Nombre'] == p['Nombre']:
                            if not (req_one_day['Inicio'] > r['Fin'] or req_one_day['Fin'] < r['Inicio']): overlap = True
                    
                    if not overlap:
                        book_request(req_one_day, occupation_map, base_schedule_turn, roster_df)
                        final_requests.append(req_one_day)
                        credits_got += 1
                        natural_days_got += 1

    final_requests = force_balance_credits(final_requests, roster_df, base_schedule_turn)
    return final_requests

# -------------------------------------------------------------------
# 3. EXPORTADORES
# -------------------------------------------------------------------

def generate_proposal_report(proposal_data, valid_requests, roster_df):
    wb = Workbook()
    fill_orange = PatternFill("solid", fgColor="FFD966")
    fill_green = PatternFill("solid", fgColor="C6EFCE")
    font_bold = Font(bold=True)
    ws1 = wb.active; ws1.title = "Comparativa de Ajustes"
    headers = ["Nombre", "Inicio Original", "Fin Original", "Inicio Propuesto", "Fin Propuesto", "Estado", "Motivo"]
    ws1.append(headers)
    for col in range(1, 8): ws1.cell(1, col).font = font_bold
    for p in proposal_data:
        row = [p['Nombre'], p['Orig_Inicio'].strftime("%d/%m/%Y"), p['Orig_Fin'].strftime("%d/%m/%Y"), p['New_Inicio'].strftime("%d/%m/%Y"), p['New_Fin'].strftime("%d/%m/%Y"), p['Status'], p['Reason']]
        ws1.append(row)
        curr = ws1.max_row
        if p['Status'] == "Modificado":
            for c in range(1, 8): ws1.cell(curr, c).fill = fill_orange
        elif p['Status'] == "Aceptado":
            for c in range(1, 8): ws1.cell(curr, c).fill = fill_green
    ws1.column_dimensions['A'].width = 20; ws1.column_dimensions['G'].width = 40
    ws2 = wb.create_sheet("Datos Listos para Subir")
    h_headers = ["ID_Puesto", "Nombre"]
    for i in range(1, 21): h_headers.extend([f"Inicio {i}", f"Fin {i}"])
    ws2.append(h_headers)
    sorted_reqs = sorted(valid_requests, key=lambda x: x['Nombre'])
    for _, person in roster_df.iterrows():
        p_name = person['Nombre']; p_id = person['ID_Puesto']
        p_reqs = [r for r in sorted_reqs if r['Nombre'] == p_name]
        p_reqs.sort(key=lambda x: x['Inicio'])
        row_data = [p_id, p_name]
        for r in p_reqs: row_data.append(r['Inicio']); row_data.append(r['Fin'])
        ws2.append(row_data)
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out

def generate_clean_import_excel(valid_requests, roster_df):
    wb = Workbook()
    ws2 = wb.active; ws2.title = "Datos Listos para Subir"
    h_headers = ["ID_Puesto", "Nombre"]
    for i in range(1, 21): h_headers.extend([f"Inicio {i}", f"Fin {i}"])
    ws2.append(h_headers)
    sorted_reqs = sorted(valid_requests, key=lambda x: x['Nombre'])
    for _, person in roster_df.iterrows():
        p_name = person['Nombre']; p_id = person['ID_Puesto']
        p_reqs = [r for r in sorted_reqs if r['Nombre'] == p_name]
        p_reqs.sort(key=lambda x: x['Inicio'])
        row_data = [p_id, p_name]
        for r in p_reqs: row_data.append(r['Inicio']); row_data.append(r['Fin'])
        ws2.append(row_data)
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out

def smart_repair_requests(roster_df, imported_requests, year, night_periods):
    base_schedule_turn, total_days = generate_base_schedule(year)
    occupation_map = {i: [] for i in range(total_days)}
    proposal_data = []
    final_valid_requests = []
    imported_requests.sort(key=lambda x: x['Nombre'])

    for req in imported_requests:
        conflict = check_request_conflict(req, occupation_map, base_schedule_turn, roster_df, night_periods, total_days)
        
        if not conflict:
            book_request(req, occupation_map, base_schedule_turn, roster_df)
            final_valid_requests.append(req)
            proposal_data.append({
                "Nombre": req['Nombre'],
                "Orig_Inicio": req['Inicio'], "Orig_Fin": req['Fin'],
                "New_Inicio": req['Inicio'], "New_Fin": req['Fin'],
                "Status": "Aceptado", "Reason": "OK"
            })
        else:
            original_start = req['Inicio']; duration = (req['Fin'] - req['Inicio']).days + 1
            found_fix = False
            shifts = []
            for i in range(1, 16): shifts.append(i); shifts.append(-i)
            
            for delta in shifts:
                new_start = original_start + datetime.timedelta(days=delta)
                new_end = new_start + datetime.timedelta(days=duration-1)
                if new_start.year != year or new_end.year != year: continue
                
                new_req = {"Nombre": req['Nombre'], "Inicio": new_start, "Fin": new_end}
                new_conflict = check_request_conflict(new_req, occupation_map, base_schedule_turn, roster_df, night_periods, total_days)
                
                if not new_conflict:
                    book_request(new_req, occupation_map, base_schedule_turn, roster_df)
                    final_valid_requests.append(new_req)
                    proposal_data.append({
                        "Nombre": req['Nombre'],
                        "Orig_Inicio": req['Inicio'], "Orig_Fin": req['Fin'],
                        "New_Inicio": new_start, "New_Fin": new_end,
                        "Status": "Modificado", "Reason": f"Conflicto original: {conflict}"
                    })
                    found_fix = True
                    break
            
            if not found_fix:
                proposal_data.append({
                    "Nombre": req['Nombre'],
                    "Orig_Inicio": req['Inicio'], "Orig_Fin": req['Fin'],
                    "New_Inicio": req['Inicio'], "New_Fin": req['Fin'],
                    "Status": "Rechazado", "Reason": "Imposible encajar +/- 15 días"
                })

    return final_valid_requests, proposal_data

def generate_visual_error_report(schedule, roster_df, year, night_periods, error_heatmap, text_errors):
    wb = Workbook()
    fill_red = PatternFill("solid", fgColor="FF0000"); fill_yellow = PatternFill("solid", fgColor="FFD700")
    fill_orange = PatternFill("solid", fgColor="FFA500"); fill_purple = PatternFill("solid", fgColor="CBC3E3")
    font_bold = Font(bold=True); font_white = Font(color="FFFFFF", bold=True)
    align_c = Alignment(horizontal="center", vertical="center")
    border_all = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws1 = wb.active; ws1.title = "Mapa de Conflictos"
    ws1.column_dimensions['A'].width = 15
    for i in range(2, 34): ws1.column_dimensions[get_column_letter(i)].width = 4
    current_row = 1
    for t in TEAMS:
        ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=32)
        cell_title = ws1.cell(current_row, 1, f"TURNO {t}"); cell_title.font = Font(bold=True, size=14, color="FFFFFF"); cell_title.fill = PatternFill("solid", fgColor="B22222"); cell_title.alignment = align_c
        current_row += 2
        team_members = roster_df[roster_df['Turno'] == t]
        for _, p in team_members.iterrows():
            name = p['Nombre']; ws1.cell(current_row, 1, name).font = font_bold
            for d in range(1, 32): c = ws1.cell(current_row, d+1, d); c.alignment=align_c; c.border=border_all
            current_row += 1
            for m_idx, mes in enumerate(MESES):
                month_num = m_idx + 1
                ws1.cell(current_row, 1, mes).font = font_bold
                days_in_month = calendar.monthrange(year, month_num)[1]
                for d in range(1, 32):
                    cell = ws1.cell(current_row, d+1); cell.border=border_all; cell.alignment=align_c
                    if d <= days_in_month:
                        date_obj = datetime.date(year, month_num, d)
                        d_idx = date_obj.timetuple().tm_yday - 1
                        status = schedule[name][d_idx]
                        val = ""; fill = PatternFill("solid", fgColor="F2F2F2")
                        if status == 'T': val = "T"; fill = PatternFill("solid", fgColor="C6EFCE")
                        elif 'V' in status: val = "V"; fill = PatternFill("solid", fgColor="FFEB9C")
                        if is_in_night_period(d_idx, year, night_periods): fill = PatternFill("solid", fgColor="A6A6A6")
                        if (name, d_idx) in error_heatmap:
                            fill = fill_red; val = error_heatmap[(name, d_idx)]
                            if val == "ERR: Turno": fill = fill_yellow
                            if val == "ERR: Max 2T": fill = fill_orange
                            cell.font = font_white
                        cell.value = val; cell.fill = fill
                    else: cell.fill = PatternFill("solid", fgColor="808080")
                current_row += 1
            current_row += 2
    ws2 = wb.create_sheet("Lista de Errores"); ws2.column_dimensions['A'].width = 80
    ws2.append(["Descripción del Conflicto"])
    for err in text_errors: ws2.append([err])
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out

def generate_error_report(df_original, errors_dict):
    wb = Workbook()
    fill_red = PatternFill("solid", fgColor="FFC7CE"); font_red = Font(color="9C0006")
    ws1 = wb.active; ws1.title = "Datos con Errores"
    headers = list(df_original.columns) + ["ERROR"]
    ws1.append(headers)
    for idx, row in df_original.iterrows():
        row_data = row.tolist()
        if idx in errors_dict:
            row_data.append(errors_dict[idx])
            ws1.append(row_data); current_row = ws1.max_row
            for col in range(1, len(row_data) + 1):
                cell = ws1.cell(row=current_row, column=col); cell.fill = fill_red; cell.font = font_red
        else: row_data.append("OK"); ws1.append(row_data)
    ws2 = wb.create_sheet("Log"); ws2.append(["Fila", "Error"])
    for idx, msg in errors_dict.items(): ws2.append([f"Fila {idx + 2}", msg])
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out

def create_final_excel(schedule, roster_df, year, requests, fill_log, counters, night_periods, adjustments_log):
    wb = Workbook()
    s_T = PatternFill("solid", fgColor="C6EFCE"); s_V = PatternFill("solid", fgColor="FFEB9C")
    s_VR = PatternFill("solid", fgColor="FFFFE0"); s_Cov = PatternFill("solid", fgColor="FFC7CE")
    s_L = PatternFill("solid", fgColor="F2F2F2"); s_Night = PatternFill("solid", fgColor="A6A6A6")
    font_bold = Font(bold=True); font_red = Font(color="9C0006", bold=True)
    align_c = Alignment(horizontal="center", vertical="center")
    border_thin = Side(border_style="thin", color="000000")
    border_all = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    ws1 = wb.active; ws1.title = "Cuadrante"
    ws1.column_dimensions['A'].width = 15
    for i in range(2, 34): ws1.column_dimensions[get_column_letter(i)].width = 4
    current_row = 1
    for t in TEAMS:
        ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=32)
        cell_title = ws1.cell(current_row, 1, f"TURNO {t}"); cell_title.font = Font(bold=True, size=14, color="FFFFFF"); cell_title.fill = PatternFill("solid", fgColor="000080"); cell_title.alignment = align_c
        current_row += 2
        team_members = roster_df[roster_df['Turno'] == t]
        for _, p in team_members.iterrows():
            name = p['Nombre']; role = p['Rol']
            ws1.cell(current_row, 1, f"{name} ({role})").font = font_bold
            for d in range(1, 32): c = ws1.cell(current_row, d+1, d); c.alignment = align_c; c.font = font_bold; c.border = border_all; c.fill = PatternFill("solid", fgColor="E0E0E0")
            current_row += 1
            for m_idx, mes in enumerate(MESES):
                month_num = m_idx + 1
                ws1.cell(current_row, 1, mes).font = font_bold; ws1.cell(current_row, 1).border = border_all
                days_in_month = calendar.monthrange(year, month_num)[1]
                for d in range(1, 32):
                    cell = ws1.cell(current_row, d+1); cell.border = border_all; cell.alignment = align_c
                    if d <= days_in_month:
                        date_obj = datetime.date(year, month_num, d)
                        day_of_year = date_obj.timetuple().tm_yday - 1
                        status = schedule[name][day_of_year]
                        val = ""; fill = s_L 
                        if status == 'T': val = "T"; fill = s_T
                        elif status == 'V': val = "V"; fill = s_V
                        elif status == 'V(L)' or status == 'V(R)': val = "v"; fill = s_VR
                        elif status.startswith('T*'):
                            covered_name = status.split('(')[1][:-1]
                            covered_p = roster_df[roster_df['Nombre'] == covered_name]
                            abbr = "?"
                            if not covered_p.empty:
                                c_role = covered_p.iloc[0]['ID_Puesto']
                                c_turn = covered_p.iloc[0]['Turno']
                                if "Subjefe" in c_role: abbr = f"S{c_turn}"
                                elif "Jefe" in c_role: abbr = f"J{c_turn}"
                                elif "Cond" in c_role: abbr = f"C{c_turn}"
                                elif "Bombero" in c_role: abbr = c_role.split()[-1]
                                else: abbr = f"?{c_turn}"
                            val = abbr; fill = s_Cov; cell.font = font_red
                        if is_in_night_period(day_of_year, year, night_periods): fill = s_Night
                        cell.value = val; cell.fill = fill
                    else: cell.fill = PatternFill("solid", fgColor="808080")
                current_row += 1
            current_row += 2 
    
    ws2 = wb.create_sheet("Estadísticas")
    ws2.column_dimensions['A'].width = 20
    headers = ["Nombre", "Turno", "Puesto", "Gastado (T)", "Coberturas (T*)", "Total Días (T+T*)", "Noches Trab.", "Total Vacs (Nat)"]
    ws2.append(headers)
    for _, p in roster_df.iterrows():
        name = p['Nombre']; sch = schedule[name]
        base_sch_turn, _ = generate_base_schedule(year)
        original_ts = base_sch_turn[p['Turno']].count('T')
        v_credits = sch.count('V'); t_cover = counters[name]
        total_work = (original_ts - v_credits) + t_cover
        v_natural = sch.count('V') + sch.count('V(L)') + sch.count('V(R)')
        nights_worked = 0
        for d_idx, s in enumerate(sch):
            is_working = s == 'T' or s.startswith('T*')
            if is_working and is_in_night_period(d_idx, year, night_periods): nights_worked += 1
        ws2.append([name, p['Turno'], p['Rol'], v_credits, t_cover, total_work, nights_worked, v_natural])

    ws3 = wb.create_sheet("Resumen Solicitudes")
    ws3.append(["Nombre", "Turno", "Rol", "Periodos Solicitados", "Días Relleno (Automático)"])
    for _, p in roster_df.iterrows():
        name = p['Nombre']
        person_reqs = [f"{r['Inicio'].strftime('%d/%m')} al {r['Fin'].strftime('%d/%m')}" for r in requests if r['Nombre'] == name]
        req_str = " | ".join(person_reqs) if person_reqs else "Sin solicitudes"
        fill_dates = fill_log[name]; fill_str = "Ninguno"
        if fill_dates:
            date_ranges = []; fill_dates.sort(); range_start = fill_dates[0]; range_end = fill_dates[0]
            for i in range(1, len(fill_dates)):
                if (fill_dates[i] - fill_dates[i-1]).days == 1: range_end = fill_dates[i]
                else:
                    if range_start == range_end: date_ranges.append(range_start.strftime('%d/%m'))
                    else: date_ranges.append(f"{range_start.strftime('%d/%m')}-{range_end.strftime('%d/%m')}")
                    range_start = fill_dates[i]; range_end = fill_dates[i]
            if range_start == range_end: date_ranges.append(range_start.strftime('%d/%m'))
            else: date_ranges.append(f"{range_start.strftime('%d/%m')}-{range_end.strftime('%d/%m')}")
            fill_str = ", ".join(date_ranges)
        ws3.append([name, p['Turno'], p['Rol'], req_str, fill_str])

    ws4 = wb.create_sheet("Ajustes de Vacaciones")
    ws4.append(["Fecha", "Trabajador (Cubre)", "Cubre a (Ausente)", "Puesto Ausente"])
    adjustments_log.sort(key=lambda x: x[0])
    for day_idx, coverer, missing in adjustments_log:
        date_obj = datetime.date(year, 1, 1) + datetime.timedelta(days=day_idx)
        missing_p = roster_df[roster_df['Nombre'] == missing]
        if not missing_p.empty: missing_role = missing_p.iloc[0]['ID_Puesto']
        else: missing_role = "Desconocido"
        ws4.append([date_obj.strftime("%d/%m/%Y"), coverer, missing, missing_role])
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out

# -------------------------------------------------------------------
# INTERFAZ STREAMLIT
# -------------------------------------------------------------------

st.set_page_config(layout="wide", page_title="Gestor V7.5")

st.title("🚒 Gestor Integral V7.5")

# 1. CONFIGURACIÓN
c1, c2 = st.columns([2, 1])
with c1:
    with st.expander("1. Configuración de Plantilla", expanded=False):
        if 'roster_data' not in st.session_state:
            st.session_state.roster_data = pd.DataFrame(DEFAULT_ROSTER)
        edited_df = st.data_editor(
            st.session_state.roster_data,
            column_config={
                "ID_Puesto": st.column_config.TextColumn(disabled=True),
                "Turno": st.column_config.SelectboxColumn(options=TEAMS, required=True),
                "Rol": st.column_config.SelectboxColumn(options=ROLES, required=True),
                "SV": st.column_config.CheckboxColumn(label="¿Es SV?", help="Puede cubrir conductor")
            },
            num_rows="dynamic",
            use_container_width=True
        )
        st.session_state.roster_data = edited_df

with c2:
    with st.expander("🌑 Periodos Nocturnos", expanded=True):
        if 'nights' not in st.session_state: st.session_state.nights = []
        
        c_dn1, c_dn2 = st.columns(2)
        dn_start = c_dn1.date_input("Inicio", value=None, label_visibility="collapsed")
        dn_end = c_dn2.date_input("Fin", value=None, label_visibility="collapsed")
        if st.button("Añadir Periodo"):
            if dn_start and dn_end: st.session_state.nights.append((dn_start, dn_end))
        
        uploaded_n = st.file_uploader("Sube Excel Nocturnas", type=['xlsx'], key="n_up", label_visibility="collapsed")
        if uploaded_n and st.button("Procesar Nocturnas"):
            try:
                df_n = pd.read_excel(uploaded_n)
                added = 0
                for _, row in df_n.iterrows():
                    val_s = row.get('Inicio') if 'Inicio' in row else row.iloc[0]
                    val_e = row.get('Fin') if 'Fin' in row else row.iloc[1]
                    if pd.isnull(val_s) or pd.isnull(val_e): continue
                    try:
                        d_s = pd.to_datetime(val_s, dayfirst=True).date()
                        d_e = pd.to_datetime(val_e, dayfirst=True).date()
                        if d_s <= d_e:
                            st.session_state.nights.append((d_s, d_e))
                            added += 1
                    except: pass 
                st.success(f"Añadidos {added} periodos.")
                st.rerun()
            except Exception as e: st.error(f"Error: {e}")

        with st.container(height=200):
            if st.session_state.nights:
                for i, (s, e) in enumerate(st.session_state.nights):
                    col_del, col_tx = st.columns([1,5])
                    if col_del.button("x", key=f"n_{i}"):
                        st.session_state.nights.pop(i)
                        st.rerun()
                    col_tx.caption(f"{s.strftime('%d/%m')} - {e.strftime('%d/%m')}")
            else:
                st.caption("Sin periodos.")

# 2. GESTOR
st.divider()
col_main, col_list = st.columns([2, 1])
names_list = edited_df['Nombre'].tolist()
today = datetime.date.today()
year_val = st.number_input("Año", value=today.year + 1)

if 'requests' not in st.session_state: st.session_state.requests = []
credits_map = calculate_spent_credits(edited_df, st.session_state.requests, year_val)

# INICIALIZAR ESTADOS
if 'pending_proposal' not in st.session_state: st.session_state.pending_proposal = None
if 'proposal_data' not in st.session_state: st.session_state.proposal_data = None
if 'error_report_data' not in st.session_state: st.session_state.error_report_data = None

with col_main:
    # --- IA SOLVER V7.4 ---
    with st.expander("🤖 Auto-Solver & Negociador (IA)", expanded=True):
        st.info("Sube tus vacaciones. La IA arregla conflictos y rellena imitando tu estilo.")
        uploaded_solver = st.file_uploader("Sube Excel Vacaciones", type=['xlsx'], key="solver_up")
        
        if uploaded_solver and st.button("✨ Analizar y Generar Propuesta"):
            try:
                df_up = pd.read_excel(uploaded_solver)
                imported_reqs = []
                for _, row in df_up.iterrows():
                    target_name = None
                    if 'ID_Puesto' in row: 
                         m = edited_df[edited_df['ID_Puesto'] == row['ID_Puesto']]
                         if not m.empty: target_name = m.iloc[0]['Nombre']
                    if not target_name and 'Nombre' in row:
                         if row['Nombre'] in names_list: target_name = row['Nombre']
                     
                    if target_name:
                        for i in range(1, 21):
                            col_s = f"Inicio {i}"; col_e = f"Fin {i}"
                            if col_s in row and col_e in row and not pd.isnull(row[col_s]):
                                try:
                                    s = pd.to_datetime(row[col_s], dayfirst=True).date()
                                    e = pd.to_datetime(row[col_e], dayfirst=True).date()
                                    imported_reqs.append({"Nombre": target_name, "Inicio": s, "Fin": e})
                                except: pass
                
                with st.spinner("Detectando patrones y ajustando..."):
                    fixed_reqs, proposal_data = smart_repair_requests(edited_df, imported_reqs, year_val, st.session_state.nights)
                    final_reqs = run_auto_solver_fill(edited_df, year_val, st.session_state.nights, fixed_reqs)
                
                st.session_state.pending_proposal = final_reqs
                st.session_state.proposal_data = proposal_data
                st.success("¡Propuesta lista!")
            except Exception as ex: st.error(f"Error: {ex}")

        if st.session_state.pending_proposal:
            st.markdown("---")
            st.write("### 🎯 Acciones sobre la Propuesta")
            c_accept, c_dl1, c_dl2 = st.columns(3)
            if c_accept.button("✅ Aceptar Propuesta y Cargar", type="primary"):
                st.session_state.requests = st.session_state.pending_proposal
                st.session_state.pending_proposal = None
                st.session_state.proposal_data = None
                st.success("¡Cargado!")
                st.rerun()
            excel_negotiation = generate_proposal_report(st.session_state.proposal_data, st.session_state.pending_proposal, edited_df)
            excel_clean = generate_clean_import_excel(st.session_state.pending_proposal, edited_df)
            c_dl1.download_button("📄 Descargar Informe Negociación", excel_negotiation, "Informe_Negociacion.xlsx")
            c_dl2.download_button("💾 Descargar Fichero Limpio", excel_clean, "Importacion_Limpia.xlsx")

    with st.expander("📂 Carga Masiva Horizontal"):
        template_df = edited_df[['ID_Puesto', 'Nombre']].copy()
        for i in range(1, 21): template_df[f'Inicio {i}'] = ""; template_df[f'Fin {i}'] = ""
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer: template_df.to_excel(writer, index=False)
        st.download_button("⬇️ Descargar Plantilla", buffer.getvalue(), "plantilla_h.xlsx")
        
        uploaded_file = st.file_uploader("Sube Excel", type=['xlsx'])
        if uploaded_file and st.button("Procesar Archivo"):
            try:
                df_upload = pd.read_excel(uploaded_file)
                count = 0; errors_found = {}; valid_requests = []
                for idx, row in df_upload.iterrows():
                    target_name = None
                    if 'ID_Puesto' in row and not pd.isnull(row['ID_Puesto']):
                        match = edited_df[edited_df['ID_Puesto'] == row['ID_Puesto']]
                        if not match.empty: target_name = match.iloc[0]['Nombre']
                    if not target_name and 'Nombre' in row:
                        if row['Nombre'] in names_list: target_name = row['Nombre']
                    if not target_name:
                        errors_found[idx] = "Trabajador no encontrado"; continue
                    row_has_error = False; row_error_msg = []; temp_reqs = []
                    for i in range(1, 21):
                        col_start = f'Inicio {i}'; col_end = f'Fin {i}'
                        if col_start in row and col_end in row:
                            val_start = row[col_start]; val_end = row[col_end]
                            if not pd.isnull(val_start) and not pd.isnull(val_end):
                                try:
                                    d_s = pd.to_datetime(val_start, dayfirst=True).date(); d_e = pd.to_datetime(val_end, dayfirst=True).date()
                                    if d_s > d_e: row_has_error = True; row_error_msg.append(f"P{i}: Fin < Inicio")
                                    elif is_night_restricted(d_s, st.session_state.nights) or is_night_restricted(d_e, st.session_state.nights): row_has_error = True; row_error_msg.append(f"P{i}: Choque Noche")
                                    else: temp_reqs.append({"Nombre": target_name, "Inicio": d_s, "Fin": d_e})
                                except: row_has_error = True; row_error_msg.append(f"P{i}: Fecha Mal")
                    if row_has_error: errors_found[idx] = "; ".join(row_error_msg)
                    else: valid_requests.extend(temp_reqs); count += 1
                if errors_found:
                    st.error(f"⛔ Errores en {len(errors_found)} filas."); st.session_state.error_report_data = generate_error_report(df_upload, errors_found)
                else:
                    st.session_state.error_report_data = None; st.session_state.requests.extend(valid_requests); st.success(f"✅ Importados {len(valid_requests)} periodos."); st.rerun()
            except Exception as e: st.error(f"Error: {e}")
        if st.session_state.error_report_data: st.download_button("📥 Descargar Informe de Errores", st.session_state.error_report_data, "Errores_Vacaciones.xlsx")

    st.subheader("2. Añadir Solicitud Manual")
    sel_name = st.selectbox("Trabajador", names_list)
    if sel_name:
        spent = credits_map.get(sel_name, 0)
        st.progress(min(spent/13, 1.0), text=f"Créditos T: {spent} / 13")
        row_p = edited_df[edited_df['Nombre'] == sel_name].iloc[0]
        base_sch, _ = generate_base_schedule(year_val)
        my_sch = base_sch[row_p['Turno']]
        view_months = list(range(1, 13))
        html_cal = "<div style='display:flex; flex-wrap:wrap; gap:5px; margin-bottom:10px;'>"
        for m in view_months:
            html_cal += f"<div style='border:1px solid #ddd; padding:2px; border-radius:3px; width:100px;'><strong>{MESES[m-1]}</strong>"
            days_in_m = calendar.monthrange(year_val, m)[1]
            html_cal += "<div style='display:grid; grid-template-columns:repeat(7, 1fr); gap:1px; font-size:9px; text-align:center;'>"
            for d in range(1, days_in_m + 1):
                dt = datetime.date(year_val, m, d); d_idx = dt.timetuple().tm_yday - 1
                status = my_sch[d_idx]; color = "#C6EFCE" if status == 'T' else "#F2F2F2"
                border = "2px solid #555" if is_in_night_period(d_idx, year_val, st.session_state.nights) else "1px solid #eee"
                html_cal += f"<div style='background-color:{color}; padding:1px; border:{border}'>{d}</div>"
            html_cal += "</div></div>"
        html_cal += "</div>"; st.markdown(html_cal, unsafe_allow_html=True)

    d_range = st.date_input("Selecciona Rango", [], help="Inicio - Fin")
    if st.button("Añadir Periodo", use_container_width=True):
        if len(d_range) == 2:
            start, end = d_range; conflict = False
            if is_night_restricted(start, st.session_state.nights) or is_night_restricted(end, st.session_state.nights): st.error("⛔ Conflicto periodo nocturno."); conflict = True
            if not conflict: st.session_state.requests.append({"Nombre": sel_name, "Inicio": start, "Fin": end}); st.success(f"Añadido: {sel_name}"); st.rerun()
        else: st.warning("Selecciona fechas.")

with col_list:
    st.subheader("Listado Solicitudes")
    if st.session_state.requests:
        indexed_requests = []; 
        for i, r in enumerate(st.session_state.requests): r_with_index = r.copy(); r_with_index['idx'] = i; indexed_requests.append(r_with_index)
        indexed_requests.sort(key=lambda x: x['Nombre'])
        grouped_reqs = {}
        for key, group in groupby(indexed_requests, lambda x: x['Nombre']): grouped_reqs[key] = list(group)
        for name, reqs in grouped_reqs.items():
            with st.expander(f"{name} ({len(reqs)})"):
                for r in reqs:
                    c_txt, c_btn = st.columns([4, 1])
                    c_txt.caption(f"{r['Inicio'].strftime('%d/%m')} - {r['Fin'].strftime('%d/%m')}")
                    if c_btn.button("🗑️", key=f"del_{r['idx']}"): st.session_state.requests.pop(r['idx']); st.rerun()
    else: st.info("Sin solicitudes.")
    if st.button("🗑️ Borrar TODO", type="secondary"): st.session_state.requests = []; st.rerun()

st.divider()
if st.button("🚀 Generar Excel Final", type="primary", use_container_width=True):
    if not st.session_state.requests: st.error("Faltan solicitudes.")
    else:
        final_sch, errs, counters, fill_log, adjustments_log, err_heatmap = validate_and_generate(edited_df, st.session_state.requests, year_val, st.session_state.nights)
        if errs:
            st.error("❌ Conflictos en solicitudes actuales."); 
            for e in errs: st.write(f"- {e}")
            error_excel = generate_visual_error_report(final_sch, edited_df, year_val, st.session_state.nights, err_heatmap, errs)
            st.download_button("📥 Descargar Mapa de Conflictos (Excel Rojo)", error_excel, "Conflictos_Visuales.xlsx")
        else:
            st.success("✅ Éxito"); excel_data = create_final_excel(final_sch, edited_df, year_val, st.session_state.requests, fill_log, counters, st.session_state.nights, adjustments_log)
            st.download_button("📥 Descargar", excel_data, f"Cuadrante_V7.5_{year_val}.xlsx")
