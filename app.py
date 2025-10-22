import streamlit as st
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import datetime
import io
import sys
import calendar 

# --- Constantes ---
TEAMS = ['A', 'B', 'C']
MESES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", 
         "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]


# -------------------------------------------------------------------
# LÓGICA DE GENERACIÓN (T2-L2)
# -------------------------------------------------------------------

def build_vacation_map(vacations_list):
    """
    Construye el mapa de vacaciones desde la lista generada por la app.
    Valida solapamientos y días totales.
    """
    v_map = {}
    day_set = set()
    
    for team, start, end in vacations_list:
        if (end - start + 1) != 13:
            st.error(f"Error interno: El periodo de {team} ({start}-{end}) no es de 13 días.")
            return None
            
        for day in range(start, end + 1):
            if day in day_set:
                st.error(f"¡Error! Vacaciones solapadas en el día {day}. Revisa las fechas.")
                return None 
            v_map[day] = team
            day_set.add(day)
            
    for team in TEAMS:
        count = sum(1 for t in v_map.values() if t == team)
        if count != 39:
            st.error(f"Error: El Turno {team} no tiene 39 días de V (tiene {count}).")
            return None
            
    return v_map

def generate_schedule(vacation_map, total_days):
    """
    Genera el cuadrante para 365 o 366 días (Lógica T2-L2).
    """
    schedule = {team: [] for team in TEAMS}
    internal_status = {'A': 'L2', 'B': 'L1', 'C': 'T'}

    for day in range(1, total_days + 1):
        
        team_on_v_today = vacation_map.get(day)
        team_on_v_yesterday = vacation_map.get(day - 1)
        
        new_status = {}

        if team_on_v_today:
            # --- ESTADO: Cobertura 2/2 ---
            v_team = team_on_v_today
            new_status[v_team] = 'V'
            active_teams = [t for t in TEAMS if t != v_team]
            
            team_1 = active_teams[0]
            team_2 = active_teams[1]
            prev_1 = internal_status[team_1]
            prev_2 = internal_status[team_2]

            if not team_on_v_yesterday:
                # Caso B: Transición de Base 1/2 -> Cobertura 2/2
                if prev_1 == 'T': 
                    new_status[team_1] = 'L1' 
                    new_status[team_2] = 'T1'
                elif prev_2 == 'T': 
                    new_status[team_1] = 'T1'
                    new_status[team_2] = 'L1'
                else:
                    if prev_1 == 'L2':
                        new_status[team_1] = 'T1'
                        new_status[team_2] = 'L1'
                    else: # prev_1=L1, prev_2=L2
                        new_status[team_1] = 'L1'
                        new_status[team_2] = 'T1'
            
            else:
                # Caso B (Continuación): Seguir en Cobertura 2/2
                for team in active_teams:
                    prev = internal_status[team]
                    if prev == 'L2': new_status[team] = 'T1'
                    elif prev == 'T1': new_status[team] = 'T2'
                    elif prev == 'T2': new_status[team] = 'L1'
                    elif prev == 'L1': new_status[team] = 'L2'

        else:
            # --- ESTADO: Base 1/2 ---
            if team_on_v_yesterday:
                # Caso C: Transición de Cobertura 2/2 -> Base 1/2
                returning_team = team_on_v_yesterday
                resting_teams = [t for t in TEAMS if t != returning_team]
                new_status[returning_team] = 'T'
                team_A = resting_teams[0]
                team_B = resting_teams[1]
                
                if internal_status[team_A].startswith('T'):
                    new_status[team_A] = 'L1'
                    new_status[team_B] = 'L2'
                else: 
                    new_status[team_A] = 'L2'
                    new_status[team_B] = 'L1'
            else:
                # --- Caso A (Continuación) ---
                # Lógica para la rotación 1/2 normal
                for team in TEAMS:
                    prev = internal_status[team]
                    if prev == 'L2':
                        new_status[team] = 'T'
                    elif prev == 'T':
                        new_status[team] = 'L1'
                    elif prev == 'L1':
                        new_status[team] = 'L2'
        
        # --- Actualización de estado ---
        # Actualiza el estado para el bucle del DÍA SIGUIENTE
        internal_status = new_status
        
        # Guarda el estado del DÍA ACTUAL en el horario
        for team in TEAMS:
            state = internal_status[team]
            if state.startswith('T'):
                schedule[team].append('T')
            elif state.startswith('L'):
                schedule[team].append('L')
            elif state == 'V':
                schedule[team].append('V')
                
    # Validación final
    for day_index in range(total_days):
        workers_today = sum(1 for team in TEAMS if schedule[team][day_index] == 'T')
        if workers_today != 1:
            st.error(f"¡Error Crítico de Lógica! Día {day_index + 1} tiene {workers_today} trabajadores.")
            return None

    return schedule

# -------------------------------------------------------------------
# FUNCIÓN DE EXCEL (FORMATO CALENDARIO)
# -------------------------------------------------------------------

def create_calendar_xlsx_in_memory(schedule, names, year):
    """
    Crea el archivo Excel en memoria con formato de matriz de calendario,
    colores y pestañas individuales.
    """
    wb = Workbook()
    
    # --- Definición de Estilos ---
    STYLE_T = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Verde
    STYLE_L = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # Gris
    STYLE_V = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid") # Azul
    STYLE_EMPTY = PatternFill(start_color="808080", end_color="808080", fill_type="solid") # Gris oscuro
    
    FONT_BLACK = Font(color="000000")
    FONT_BOLD = Font(bold=True)
    ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
    
    styles = {'T': STYLE_T, 'L': STYLE_L, 'V': STYLE_V}

    def set_calendar_widths(ws):
        # Ancho para la columna de Mes
        ws.column_dimensions['A'].width = 12
        # Ancho para las columnas de días (1-31)
        for i in range(2, 33):
            ws.column_dimensions[get_column_letter(i)].width = 4

    def create_calendar_grid(ws, start_row, team_schedule, year):
        # Fila de Cabecera (Días 1-31)
        ws.cell(row=start_row, column=1, value="Mes").font = FONT_BOLD
        for d in range(1, 32):
            cell = ws.cell(row=start_row, column=d + 1, value=d)
            cell.font = FONT_BOLD
            cell.alignment = ALIGN_CENTER
        
        # Filas de Meses (Enero-Diciembre)
        current_row = start_row + 1
        for m in range(1, 13):
            ws.cell(row=current_row, column=1, value=MESES[m-1]).font = FONT_BOLD
            
            # Obtener días en el mes
            num_days_in_month = calendar.monthrange(year, m)[1]
            
            for d in range(1, 32):
                cell = ws.cell(row=current_row, column=d + 1)
                
                if d <= num_days_in_month:
                    # Es un día real, obtener estado
                    try:
                        day_of_year = datetime.date(year, m, d).timetuple().tm_yday
                        # El índice de la lista es (día del año - 1)
                        status = team_schedule[day_of_year - 1]
                        cell.value = status
                        cell.fill = styles.get(status)
                        cell.font = FONT_BLACK
                    except IndexError:
                        # Esto pasaría si el schedule tiene 365 días y el año es 366
                        cell.value = "ERR"
                else:
                    # Día no existe (ej. 30 Feb, 31 Abr)
                    cell.fill = STYLE_EMPTY
                
                cell.alignment = ALIGN_CENTER
            current_row += 1
        return current_row # Devuelve la siguiente fila libre

    # --- 1. Hoja "Cuadrante General" ---
    ws_general = wb.active
    ws_general.title = "Cuadrante General"
    
    current_row = 1
    for team_id in TEAMS: # 'A', 'B', 'C'
        team_name = names[team_id]
        team_schedule = schedule[team_id]
        
        # Título de la persona
        cell_title = ws_general.cell(row=current_row, column=1, value=team_name)
        cell_title.font = Font(bold=True, size=14)
        ws_general.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=32)
        cell_title.alignment = ALIGN_CENTER
        
        current_row += 1 # Moverse a la fila de cabecera de días
        
        # Crear la matriz para esta persona
        current_row = create_calendar_grid(ws_general, current_row, team_schedule, year)
        
        current_row += 2 # Añadir espacio antes de la siguiente persona

    set_calendar_widths(ws_general)

    # --- 2. Hojas Individuales ---
    for team_id in TEAMS: # 'A', 'B', 'C'
        team_name = names[team_id]
        team_schedule = schedule[team_id]
        # Acortar nombre de pestaña si es muy largo
        safe_team_name = team_name[:31]
        ws_ind = wb.create_sheet(title=safe_team_name)
        
        # Crear la matriz (empezando en Fila 1)
        create_calendar_grid(ws_ind, 1, team_schedule, year)
        set_calendar_widths(ws_ind)

    # --- Guardar en Memoria ---
    mem_file = io.BytesIO()
    wb.save(mem_file)
    mem_file.seek(0)
    return mem_file.getvalue()


# -------------------------------------------------------------------
# APLICACIÓN WEB (Streamlit)
# -------------------------------------------------------------------

st.set_page_config(layout="wide")
st.title("📅 Generador de Cuadrantes 3x3 (1T-2L / 2T-2L)")
st.write("""
Introduce los nombres de las 3 personas (o turnos) y las fechas de **inicio** de sus 3 periodos de vacaciones (de 13 días cada uno). 
Todas las fechas deben ser del mismo año. La aplicación detectará si es un año bisiesto.
""")

# --- Contenedor del formulario ---
with st.form("schedule_form"):

    st.header("1. Nombres de los Turnos")
    col1, col2, col3 = st.columns(3)
    with col1:
        name_a = st.text_input("Nombre Turno A", "Persona A")
    with col2:
        name_b = st.text_input("Nombre Turno B", "Persona B")
    with col3:
        name_c = st.text_input("Nombre Turno C", "Persona C")
    
    names = {'A': name_a, 'B': name_b, 'C': name_c}

    st.header("2. Periodos de Vacaciones (13 días cada uno)")
    st.warning("Asegúrate de que los 9 periodos (3 por persona) no se solapen.", icon="⚠️")
    
    today = datetime.date.today()
    default_year = today.year + 1 if today.month > 6 else today.year
    
    col1, col2, col3 = st.columns(3)
    
    # Fechas de A
    with col1:
        st.subheader(f"Vacaciones de {name_a}")
        a_v1 = st.date_input(f"Inicio Periodo 1 ({name_a})", datetime.date(default_year, 2, 1))
        a_v2 = st.date_input(f"Inicio Periodo 2 ({name_a})", datetime.date(default_year, 6, 1))
        a_v3 = st.date_input(f"Inicio Periodo 3 ({name_a})", datetime.date(default_year, 10, 1))

    # Fechas de B
    with col2:
        st.subheader(f"Vacaciones de {name_b}")
        b_v1 = st.date_input(f"Inicio Periodo 1 ({name_b})", datetime.date(default_year, 3, 1))
        b_v2 = st.date_input(f"Inicio Periodo 2 ({name_b})", datetime.date(default_year, 7, 1))
        b_v3 = st.date_input(f"Inicio Periodo 3 ({name_b})", datetime.date(default_year, 11, 1))

    # Fechas de C
    with col3:
        st.subheader(f"Vacaciones de {name_c}")
        c_v1 = st.date_input(f"Inicio Periodo 1 ({name_c})", datetime.date(default_year, 4, 1))
        c_v2 = st.date_input(f"Inicio Periodo 2 ({name_c})", datetime.date(default_year, 8, 1))
        c_v3 = st.date_input(f"Inicio Periodo 3 ({name_c})", datetime.date(default_year, 12, 1))

    st.divider()
    submitted = st.form_submit_button("Generar Cuadrante", type="primary", use_container_width=True)


# --- Lógica de procesamiento ---
if submitted:
    
    base_dates = [
        ('A', a_v1), ('A', a_v2), ('A', a_v3),
        ('B', b_v1), ('B', b_v2), ('B', b_v3),
        ('C', c_v1), ('C', c_v2), ('C', c_v3),
    ]
    
    vacations_list = []
    
    try:
        # Detectar año y días totales
        first_year = base_dates[0][1].year
        if any(d.year != first_year for _, d in base_dates):
            st.error("¡Error! Todas las fechas deben ser del mismo año.")
            sys.exit() 
        
        is_leap = calendar.isleap(first_year)
        total_days = 366 if is_leap else 365
        st.info(f"Detectado {first_year}: {total_days} días (Año {'Bisiesto' if is_leap else 'Normal'}).")

        for team, start_date in base_dates:
            end_date = start_date + datetime.timedelta(days=12)
            start_day_of_year = start_date.timetuple().tm_yday
            end_day_of_year = end_date.timetuple().tm_yday
            
            if start_day_of_year > end_day_of_year: # Cruce de año
                st.error(f"Error: El periodo de {team} que empieza el {start_date} cruza al año siguiente.")
                sys.exit()
            if end_day_of_year > total_days:
                st.error(f"Error: El periodo de {team} que empieza el {start_date} se sale del año {first_year}.")
                sys.exit()

            vacations_list.append((team, start_day_of_year, end_day_of_year))

        # 2. Construir el mapa de vacaciones (valida solapamientos)
        with st.spinner("Validando fechas y solapamientos..."):
            vacation_map = build_vacation_map(vacations_list)
        
        if vacation_map:
            # 3. Pasar total_days
            with st.spinner(f"Generando cuadrante de {total_days} días..."):
                schedule = generate_schedule(vacation_map, total_days)
            
            if schedule:
                st.success("¡Cuadrante generado con éxito!")
                
                # 4. Llamar a la nueva función de Excel y pasar el año
                excel_data = create_calendar_xlsx_in_memory(schedule, names, first_year)
                
                st.download_button(
                    label="📥 Descargar Cuadrante (Formato Calendario)",
                    data=excel_data,
                    file_name=f"cuadrante_calendario_{first_year}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

    except Exception as e:
        st.error(f"Ha ocurrido un error inesperado: {e}")