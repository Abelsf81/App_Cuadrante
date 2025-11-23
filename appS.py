import streamlit as st
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime
import io
import random
import calendar 
import pandas as pd
from itertools import groupby
from operator import itemgetter

# --- CONSTANTES Y CONFIGURACIÓN ---
TEAMS = ['A', 'B', 'C']
ROLES = ["Jefe", "Subjefe", "Conductor", "Bombero"] 
MESES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", 
         "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

# Plantilla por defecto
DEFAULT_ROSTER = [
    {"ID_Puesto": "Jefe A",       "Nombre": "Jefe A",       "Turno": "A", "Rol": "Jefe",       "SV": False},
    {"ID_Puesto": "Subjefe A",    "Nombre": "Subjefe A",    "Turno": "A", "Rol": "Subjefe",    "SV": False},
    {"ID_Puesto": "Cond A",       "Nombre": "Cond A",       "Turno": "A", "Rol": "Conductor",  "SV": True},
    {"ID_Puesto": "Bombero A1",   "Nombre": "Bombero A1",   "Turno": "A", "Rol": "Bombero",    "SV": True},
    {"ID_Puesto": "Bombero A2",   "Nombre": "Bombero A2",   "Turno": "A", "Rol": "Bombero",    "SV": False},
    {"ID_Puesto": "Bombero A3",   "Nombre": "Bombero A3",   "Turno": "A", "Rol": "Bombero",    "SV": False},
    
    {"ID_Puesto": "Jefe B",       "Nombre": "Jefe B",       "Turno": "B", "Rol": "Jefe",       "SV": False},
    {"ID_Puesto": "Subjefe B",    "Nombre": "Subjefe B",    "Turno": "B", "Rol": "Subjefe",    "SV": False},
    {"ID_Puesto": "Cond B",       "Nombre": "Cond B",       "Turno": "B", "Rol": "Conductor",  "SV": True},
    {"ID_Puesto": "Bombero B1",   "Nombre": "Bombero B1",   "Turno": "B", "Rol": "Bombero",    "SV": True},
    {"ID_Puesto": "Bombero B2",   "Nombre": "Bombero B2",   "Turno": "B", "Rol": "Bombero",    "SV": False},
    {"ID_Puesto": "Bombero B3",   "Nombre": "Bombero B3",   "Turno": "B", "Rol": "Bombero",    "SV": False},

    {"ID_Puesto": "Jefe C",       "Nombre": "Jefe C",       "Turno": "C", "Rol": "Jefe",       "SV": False},
    {"ID_Puesto": "Subjefe C",    "Nombre": "Subjefe C",    "Turno": "C", "Rol": "Subjefe",    "SV": False},
    {"ID_Puesto": "Cond C",       "Nombre": "Cond C",       "Turno": "C", "Rol": "Conductor",  "SV": True},
    {"ID_Puesto": "Bombero C1",   "Nombre": "Bombero C1",   "Turno": "C", "Rol": "Bombero",    "SV": True},
    {"ID_Puesto": "Bombero C2",   "Nombre": "Bombero C2",   "Turno": "C", "Rol": "Bombero",    "SV": False},
    {"ID_Puesto": "Bombero C3",   "Nombre": "Bombero C3",   "Turno": "C", "Rol": "Bombero",    "SV": False},
]

# -------------------------------------------------------------------
# 1. MOTOR LÓGICO
# -------------------------------------------------------------------

def generate_base_schedule(year):
    is_leap = calendar.isleap(year)
    total_days = 366 if is_leap else 365
    status = {'A': 0, 'B': 2, 'C': 1} 
    schedule = {team: [] for team in TEAMS}
    for _ in range(total_days):
        for t in TEAMS:
            if status[t] == 0: schedule[t].append('T')
            else: schedule[t].append('L')
            status[t] = (status[t] + 1) % 3
    return schedule, total_days

def get_candidates(person_missing, roster_df, day_idx, current_schedule, adjustments_log_current_day=None):
    candidates = []
    missing_role = person_missing['Rol']
    missing_turn = person_missing['Turno']
    
    blocked_turns = set()
    if adjustments_log_current_day:
        for coverer_name in adjustments_log_current_day:
            cov_p = roster_df[roster_df['Nombre'] == coverer_name]
            if not cov_p.empty:
                blocked_turns.add(cov_p.iloc[0]['Turno'])

    for _, candidate in roster_df.iterrows():
        if candidate['Turno'] == missing_turn: continue
        cand_status = current_schedule[candidate['Nombre']][day_idx]
        if cand_status != 'L': continue 
        if candidate['Turno'] in blocked_turns: continue

        is_compatible = False
        cand_role = candidate['Rol']
        
        if missing_role == "Jefe":
            if cand_role in ["Jefe", "Subjefe"]: is_compatible = True
        elif missing_role == "Subjefe":
             if cand_role in ["Jefe", "Subjefe"]: is_compatible = True
        elif missing_role == "Conductor":
            if cand_role == "Conductor": is_compatible = True
            if candidate['SV']: is_compatible = True
        elif missing_role == "Bombero":
            if cand_role == "Bombero": is_compatible = True
            if candidate['SV']: is_compatible = True
            
        if is_compatible:
            candidates.append(candidate['Nombre'])
    return candidates

def is_in_night_period(day_idx, year, night_periods):
    current_date = datetime.date(year, 1, 1) + datetime.timedelta(days=day_idx)
    for start, end in night_periods:
        if start <= current_date <= end: return True
    return False

def get_night_transition_dates(night_periods):
    dates = set()
    for start, end in night_periods:
        dates.add(start)
        dates.add(end)
    return dates

def get_working_days_count(person_name, start_date, end_date, roster_df, year):
    """Cuenta cuantos días T hay en un rango"""
    base_sch, _ = generate_base_schedule(year)
    row = roster_df[roster_df['Nombre'] == person_name]
    if row.empty: return 0
    turn = row.iloc[0]['Turno']
    
    s_idx = start_date.timetuple().tm_yday - 1
    e_idx = end_date.timetuple().tm_yday - 1
    count = 0
    for d in range(s_idx, e_idx + 1):
        if base_sch[turn][d] == 'T': count += 1
    return count

def calculate_spent_credits_interactive(roster_df, requests, year):
    """Calcula creditos y devuelve detalle por persona"""
    base_sch, _ = generate_base_schedule(year)
    stats = {}
    
    for _, p in roster_df.iterrows():
        stats[p['Nombre']] = {'credits': 0, 'natural': 0}

    for req in requests:
        name = req['Nombre']
        if name not in stats: continue
        
        s = req['Inicio']
        e = req['Fin']
        row = roster_df[roster_df['Nombre'] == name]
        if row.empty: continue
        turn = row.iloc[0]['Turno']
        
        s_idx = s.timetuple().tm_yday - 1
        e_idx = e.timetuple().tm_yday - 1
        
        days_nat = (e_idx - s_idx) + 1
        credits_t = 0
        for d in range(s_idx, e_idx + 1):
            if base_sch[turn][d] == 'T': credits_t += 1
            
        stats[name]['credits'] += credits_t
        stats[name]['natural'] += days_nat
        
    return stats

def check_conflicts_interactive(roster_df, requests, year, night_periods):
    """Analiza conflictos en tiempo real y devuelve lista"""
    base_schedule_turn, total_days = generate_base_schedule(year)
    occupation_map = {i: [] for i in range(total_days)}
    conflicts = []
    transition_dates = get_night_transition_dates(night_periods)

    # Llenar mapa
    for req in requests:
        person = roster_df[roster_df['Nombre'] == req['Nombre']].iloc[0]
        s_idx = req['Inicio'].timetuple().tm_yday - 1
        e_idx = req['Fin'].timetuple().tm_yday - 1
        
        # Chequeo Nocturna
        for d in range(s_idx, e_idx + 1):
            day_obj = datetime.date(year, 1, 1) + datetime.timedelta(days=d)
            if day_obj in transition_dates:
                if base_schedule_turn[person['Turno']][d] == 'T':
                    conflicts.append(f"⛔ {person['Nombre']}: Trabaja en Cambio de Turno ({day_obj.strftime('%d/%m')})")

            if base_schedule_turn[person['Turno']][d] == 'T':
                occupation_map[d].append(person.to_dict())

    # Analizar mapa
    for d, occupants in occupation_map.items():
        if len(occupants) > 2:
            day_obj = datetime.date(year, 1, 1) + datetime.timedelta(days=d)
            names = [o['Nombre'] for o in occupants]
            conflicts.append(f"💥 {day_obj.strftime('%d/%m')}: {len(occupants)} personas ({', '.join(names)})")
        
        if len(occupants) == 2:
            o1, o2 = occupants[0], occupants[1]
            day_obj = datetime.date(year, 1, 1) + datetime.timedelta(days=d)
            
            if o1['Turno'] == o2['Turno']:
                conflicts.append(f"⚠️ {day_obj.strftime('%d/%m')}: Mismo Turno ({o1['Nombre']} y {o2['Nombre']})")
            
            # Regla Categoría
            if o1['Rol'] == o2['Rol'] and o1['Rol'] != "Bombero":
                 conflicts.append(f"⚠️ {day_obj.strftime('%d/%m')}: Misma Categoría {o1['Rol']} ({o1['Nombre']} y {o2['Nombre']})")
    
    return list(set(conflicts)) # Unicos

# -------------------------------------------------------------------
# GENERADORES EXCEL (Simplificados para el final)
# -------------------------------------------------------------------
def validate_and_generate_final(roster_df, requests, year, night_periods):
    # Esta función es la misma lógica de validación final que tenías, 
    # pero se usa solo al dar al botón final.
    # Reutilizamos la lógica de app anterior pero simplificada porque
    # se supone que el usuario ya ha limpiado los conflictos en la UI.
    
    base_schedule_turn, total_days = generate_base_schedule(year)
    final_schedule = {} 
    turn_coverage_counters = {'A': 0, 'B': 0, 'C': 0}
    person_coverage_counters = {name: 0 for name in roster_df['Nombre']}
    name_to_turn = {row['Nombre']: row['Turno'] for _, row in roster_df.iterrows()}
    
    for _, row in roster_df.iterrows():
        final_schedule[row['Nombre']] = base_schedule_turn[row['Turno']].copy()

    day_vacations = {i: [] for i in range(total_days)}
    
    for req in requests:
        name = req['Nombre']
        s_idx = req['Inicio'].timetuple().tm_yday - 1
        e_idx = req['Fin'].timetuple().tm_yday - 1
        for d in range(s_idx, e_idx + 1):
            if final_schedule[name][d] == 'T':
                day_vacations[d].append(name)
                final_schedule[name][d] = 'V'
            else:
                final_schedule[name][d] = 'V(L)'

    adjustments_log = []
    # Solo calculamos coberturas, asumimos que conflictos graves ya se arreglaron visualmente
    # o se ignorarán si el usuario persiste.
    
    for d in range(total_days):
        absent_people = day_vacations[d]
        if not absent_people: continue
        
        current_day_coverers = []
        # Prioridad Mandos
        absent_people.sort(key=lambda x: 0 if "Jefe" in x or "Subjefe" in x else 1)

        for name_missing in absent_people:
            person_row = roster_df[roster_df['Nombre'] == name_missing].iloc[0]
            candidates = get_candidates(person_row, roster_df, d, final_schedule, current_day_coverers)
            
            if candidates:
                # Filtrar 2T
                valid = []
                for c in candidates:
                    prev = final_schedule[c][d-1] if d>0 else 'L'
                    prev2 = final_schedule[c][d-2] if d>1 else 'L'
                    if not (prev.startswith('T') and prev2.startswith('T')): valid.append(c)
                
                if valid:
                    valid.sort(key=lambda x: (turn_coverage_counters[name_to_turn[x]], person_coverage_counters[x], random.random()))
                    chosen = valid[0]
                    final_schedule[chosen][d] = f"T*({name_missing})"
                    adjustments_log.append((d, chosen, name_missing))
                    current_day_coverers.append(chosen)
                    turn_coverage_counters[name_to_turn[chosen]] += 1
                    person_coverage_counters[chosen] += 1

    fill_log = {} # Ya no rellenamos, lo hace el usuario
    return final_schedule, adjustments_log, person_coverage_counters, fill_log

def create_final_excel(schedule, roster_df, year, requests, fill_log, counters, night_periods, adjustments_log):
    wb = Workbook()
    s_T = PatternFill("solid", fgColor="C6EFCE"); s_V = PatternFill("solid", fgColor="FFEB9C")
    s_VR = PatternFill("solid", fgColor="FFFFE0"); s_Cov = PatternFill("solid", fgColor="FFC7CE")
    s_L = PatternFill("solid", fgColor="F2F2F2"); s_Night = PatternFill("solid", fgColor="A6A6A6")
    font_bold = Font(bold=True); font_red = Font(color="9C0006", bold=True)
    align_c = Alignment(horizontal="center", vertical="center")
    border_thin = Side(border_style="thin", color="000000")
    border_all = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    ws1 = wb.active; ws1.title = "Cuadrante"
    ws1.column_dimensions['A'].width = 15
    for i in range(2, 34): ws1.column_dimensions[get_column_letter(i)].width = 4
    current_row = 1
    for t in TEAMS:
        ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=32)
        cell_title = ws1.cell(current_row, 1, f"TURNO {t}"); cell_title.font = Font(bold=True, size=14, color="FFFFFF"); cell_title.fill = PatternFill("solid", fgColor="000080"); cell_title.alignment = align_c
        current_row += 2
        team_members = roster_df[roster_df['Turno'] == t]
        for _, p in team_members.iterrows():
            name = p['Nombre']; role = p['Rol']
            ws1.cell(current_row, 1, f"{name} ({role})").font = font_bold
            for d in range(1, 32): c = ws1.cell(current_row, d+1, d); c.alignment = align_c; c.font = font_bold; c.border = border_all; c.fill = PatternFill("solid", fgColor="E0E0E0")
            current_row += 1
            for m_idx, mes in enumerate(MESES):
                month_num = m_idx + 1
                ws1.cell(current_row, 1, mes).font = font_bold; ws1.cell(current_row, 1).border = border_all
                days_in_month = calendar.monthrange(year, month_num)[1]
                for d in range(1, 32):
                    cell = ws1.cell(current_row, d+1); cell.border = border_all; cell.alignment = align_c
                    if d <= days_in_month:
                        date_obj = datetime.date(year, month_num, d)
                        day_of_year = date_obj.timetuple().tm_yday - 1
                        status = schedule[name][day_of_year]
                        val = ""; fill = s_L 
                        if status == 'T': val = "T"; fill = s_T
                        elif status == 'V': val = "V"; fill = s_V
                        elif status == 'V(L)' or status == 'V(R)': val = "v"; fill = s_VR
                        elif status.startswith('T*'):
                            covered_name = status.split('(')[1][:-1]
                            covered_p = roster_df[roster_df['Nombre'] == covered_name]
                            abbr = "?"
                            if not covered_p.empty:
                                c_role = covered_p.iloc[0]['ID_Puesto']
                                c_turn = covered_p.iloc[0]['Turno']
                                if "Subjefe" in c_role: abbr = f"S{c_turn}"
                                elif "Jefe" in c_role: abbr = f"J{c_turn}"
                                elif "Cond" in c_role: abbr = f"C{c_turn}"
                                elif "Bombero" in c_role: abbr = c_role.split()[-1]
                                else: abbr = f"?{c_turn}"
                            val = abbr; fill = s_Cov; cell.font = font_red
                        if is_in_night_period(day_of_year, year, night_periods): fill = s_Night
                        cell.value = val; cell.fill = fill
                    else: cell.fill = PatternFill("solid", fgColor="808080")
                current_row += 1
            current_row += 2 
    
    ws2 = wb.create_sheet("Estadísticas")
    headers = ["Nombre", "Turno", "Puesto", "Gastado (T)", "Coberturas (T*)", "Total Vacs (Nat)"]
    ws2.append(headers)
    for _, p in roster_df.iterrows():
        name = p['Nombre']; sch = schedule[name]
        v_credits = sch.count('V'); t_cover = counters[name]
        v_natural = sch.count('V') + sch.count('V(L)') + sch.count('V(R)')
        ws2.append([name, p['Turno'], p['Rol'], v_credits, t_cover, v_natural])

    ws4 = wb.create_sheet("Ajustes de Vacaciones")
    ws4.append(["Fecha", "Trabajador (Cubre)", "Cubre a (Ausente)", "Puesto Ausente"])
    adjustments_log.sort(key=lambda x: x[0])
    for day_idx, coverer, missing in adjustments_log:
        date_obj = datetime.date(year, 1, 1) + datetime.timedelta(days=day_idx)
        missing_p = roster_df[roster_df['Nombre'] == missing]
        if not missing_p.empty: missing_role = missing_p.iloc[0]['ID_Puesto']
        else: missing_role = "Desconocido"
        ws4.append([date_obj.strftime("%d/%m/%Y"), coverer, missing, missing_role])
    
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out

# -------------------------------------------------------------------
# INTERFAZ STREAMLIT (V8.0: SIMULADOR INTERACTIVO)
# -------------------------------------------------------------------

st.set_page_config(layout="wide", page_title="Gestor V8.0 - Interactivo")

st.title("🚒 Gestor V8.0: Simulador Interactivo")
st.caption("Sube tus datos, visualiza conflictos y corrígelos en tiempo real.")

# 1. CONFIGURACIÓN INICIAL
with st.sidebar:
    st.header("1. Configuración")
    year_val = st.number_input("Año", value=2026)
    
    with st.expander("Plantilla"):
        if 'roster_data' not in st.session_state:
            st.session_state.roster_data = pd.DataFrame(DEFAULT_ROSTER)
        edited_df = st.data_editor(st.session_state.roster_data, use_container_width=True)
        st.session_state.roster_data = edited_df
        
    with st.expander("Nocturnas"):
        if 'nights' not in st.session_state: st.session_state.nights = []
        c1, c2 = st.columns(2)
        dn_s = c1.date_input("Inicio", key="n_s", value=None)
        dn_e = c2.date_input("Fin", key="n_e", value=None)
        if st.button("Añadir Nocturna"):
            if dn_s and dn_e: st.session_state.nights.append((dn_s, dn_e))
        
        st.write(f"Periodos: {len(st.session_state.nights)}")
        if st.button("Limpiar Nocturnas"): st.session_state.nights = []

# 2. ESTADO DE SOLICITUDES
if 'raw_requests_df' not in st.session_state:
    # Estructura vacía para el editor
    cols = ["Nombre", "Inicio", "Fin"]
    st.session_state.raw_requests_df = pd.DataFrame(columns=cols)

# 3. CARGA DE DATOS
st.divider()
c_up, c_info = st.columns([1, 2])

with c_up:
    uploaded_file = st.file_uploader("📂 Cargar Excel (Formato Horizontal)", type=['xlsx'])
    if uploaded_file:
        if st.button("Procesar Carga"):
            try:
                df = pd.read_excel(uploaded_file)
                new_reqs = []
                for _, row in df.iterrows():
                    name = row.get('Nombre')
                    if not name: continue
                    # Buscar pares Inicio X / Fin X
                    for i in range(1, 21):
                        ks = f"Inicio {i}"; ke = f"Fin {i}"
                        if ks in row and ke in row and not pd.isnull(row[ks]):
                            try:
                                new_reqs.append({
                                    "Nombre": name,
                                    "Inicio": pd.to_datetime(row[ks]).date(),
                                    "Fin": pd.to_datetime(row[ke]).date()
                                })
                            except: pass
                st.session_state.raw_requests_df = pd.DataFrame(new_reqs)
                st.success(f"Cargados {len(new_reqs)} registros.")
            except Exception as e: st.error(f"Error: {e}")

# 4. ZONA INTERACTIVA (EL CEREBRO)
st.subheader("2. Mesa de Trabajo Interactiva")

# Convertir DF a lista de dicts para lógica
current_requests = st.session_state.raw_requests_df.to_dict('records')

# -- ÁREA DE EDICIÓN --
c_editor, c_stats = st.columns([2, 1])

with c_editor:
    st.markdown("### ✏️ Edición de Solicitudes")
    # El usuario edita aquí. Convertimos fecha a datetime para el editor
    df_editor = st.session_state.raw_requests_df.copy()
    if not df_editor.empty:
        df_editor["Inicio"] = pd.to_datetime(df_editor["Inicio"])
        df_editor["Fin"] = pd.to_datetime(df_editor["Fin"])
    
    edited_requests_df = st.data_editor(df_editor, num_rows="dynamic", use_container_width=True, key="editor_main")
    
    # Convertir de vuelta a date para lógica
    if not edited_requests_df.empty:
        final_reqs_list = []
        for _, r in edited_requests_df.iterrows():
            try:
                final_reqs_list.append({
                    "Nombre": r["Nombre"],
                    "Inicio": r["Inicio"].date(),
                    "Fin": r["Fin"].date()
                })
            except: pass # Ignorar filas vacías o errores fecha
        current_requests = final_reqs_list

# -- LÓGICA EN TIEMPO REAL --
stats = calculate_spent_credits_interactive(edited_df, current_requests, year_val)
conflicts = check_conflicts_interactive(edited_df, current_requests, year_val, st.session_state.nights)

with c_stats:
    st.markdown("### 📊 Estado en Tiempo Real")
    
    # 1. ALERTAS DE CONFLICTOS
    if conflicts:
        st.error(f"⛔ {len(conflicts)} Conflictos Detectados")
        for c in conflicts:
            st.caption(c)
    else:
        st.success("✅ Sin conflictos normativos")

    st.divider()
    
    # 2. CONTROL DE CRÉDITOS (13)
    st.markdown("#### 🎯 Control de Créditos (13)")
    
    # Botón mágico para arreglar los de 12
    if st.button("🪄 Arreglar (Rellenar 1 día a los de 12)"):
        base_sch, total_days = generate_base_schedule(year_val)
        all_days = [datetime.date(year_val, 1, 1) + datetime.timedelta(days=i) for i in range(total_days)]
        occ_map = {i:[] for i in range(total_days)} # Mapa simple para check
        
        # Llenar mapa rapido
        for req in current_requests:
            p = edited_df[edited_df['Nombre'] == req['Nombre']].iloc[0]
            s = req['Inicio'].timetuple().tm_yday -1
            e = req['Fin'].timetuple().tm_yday -1
            for d in range(s, e+1): 
                if base_sch[p['Turno']][d] == 'T': occ_map[d].append(p['Nombre'])

        added_count = 0
        for name, data in stats.items():
            if data['credits'] == 12:
                # Buscar 1 día
                person = edited_df[edited_df['Nombre'] == name].iloc[0]
                found = False
                # Buscar un dia T libre
                random.shuffle(all_days)
                for day in all_days:
                    d_idx = day.timetuple().tm_yday - 1
                    if base_sch[person['Turno']][d_idx] == 'T':
                        # Check conflicto basico
                        is_conflict = False
                        # 1. Nocturna
                        if is_in_night_period(d_idx, year_val, st.session_state.nights): is_conflict = True
                        # 2. Ocupacion
                        occupants = occ_map[d_idx]
                        if len(occupants) >= 2: is_conflict = True
                        # 3. Mismo Turno
                        for occ_name in occupants:
                            occ_p = edited_df[edited_df['Nombre'] == occ_name].iloc[0]
                            if occ_p['Turno'] == person['Turno']: is_conflict = True
                            if occ_p['Rol'] == person['Rol'] and person['Rol'] != "Bombero": is_conflict = True
                        
                        if not is_conflict:
                            # EUREKA
                            new_req = {"Nombre": name, "Inicio": day, "Fin": day}
                            current_requests.append(new_req)
                            occ_map[d_idx].append(name)
                            added_count += 1
                            found = True
                            break
        
        if added_count > 0:
            # Actualizar estado sesión
            df_update = pd.DataFrame(current_requests)
            df_update['Inicio'] = pd.to_datetime(df_update['Inicio'])
            df_update['Fin'] = pd.to_datetime(df_update['Fin'])
            st.session_state.raw_requests_df = df_update
            st.success(f"¡Añadidos {added_count} días automáticamente! Recarga la tabla.")
            st.rerun()
        else:
            st.warning("No encontré huecos fáciles para los de 12. Tendrás que hacerlo a mano.")

    # Lista de estado
    for name, data in stats.items():
        c = data['credits']
        n = data['natural']
        
        col_st = "red" if c != 13 else "green"
        icon = "✅" if c == 13 else "⚠️"
        if c > 13: icon = "🛑"
        
        with st.expander(f"{icon} {name}: {c} Créditos ({n} Nat)", expanded=(c!=13)):
            st.progress(min(c/13, 1.0))
            if c < 13: st.caption("Faltan días de trabajo.")
            if c > 13: st.caption("Te pasaste de días.")

# 5. GENERACIÓN FINAL
st.divider()
if st.button("🚀 Generar Excel Final (Solo si todo está verde)", type="primary", use_container_width=True):
    if conflicts:
        st.error("Aún hay conflictos. Resuélvelos antes de generar.")
    else:
        sch, adj, count, fill, logs = validate_and_generate_final(edited_df, current_requests, year_val, st.session_state.nights)
        excel_io = create_final_excel(sch, edited_df, year_val, current_requests, fill, count, st.session_state.nights, adj)
        st.download_button("📥 Descargar Cuadrante Validado", excel_io, f"Cuadrante_Final_{year_val}.xlsx")
