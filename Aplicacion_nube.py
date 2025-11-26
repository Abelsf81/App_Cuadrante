import streamlit as st
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime
import io
import random
import calendar 
import pandas as pd
import os
from itertools import groupby
from operator import itemgetter
from datetime import timedelta

# ==============================================================================
# 1. CONFIGURACIÓN Y CONSTANTES
# ==============================================================================

st.set_page_config(layout="wide", page_title="Gestor V45.1")

TEAMS = ['A', 'B', 'C']
ROLES = ["Jefe", "Subjefe", "Conductor", "Bombero"] 
MESES = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
DB_FILE = "vacaciones_db.csv"
ADJ_FILE = "ajustes_db.csv"

# --- ESTRATEGIAS ---
STRATEGIES = {
    "standard": {
        "name": "🛡️ Estándar (4 Bloques)",
        "desc": "10+10+10+9 días. Requiere iniciar uno en T.",
        "blocks": [
            {"dur": 10, "cred": 4, "label": "Bloque 10d (4 Cr)"},
            {"dur": 10, "cred": 3, "label": "Bloque 10d (3 Cr)"},
            {"dur": 9,  "cred": 3, "label": "Bloque 9d (3 Cr)"}
        ],
        "auto_recipe": [ {"dur": 10, "target": 4}, {"dur": 10, "target": 3}, {"dur": 10, "target": 3}, {"dur": 9, "target": 3} ]
    },
    "safe": {
        "name": "🔢 Matemática Pura",
        "desc": "12+12+9+6 días. Indestructible.",
        "blocks": [
            {"dur": 12, "cred": 4, "label": "Largo 12d (4 Cr)"},
            {"dur": 9,  "cred": 3, "label": "Medio 9d (3 Cr)"},
            {"dur": 6,  "cred": 2, "label": "Corto 6d (2 Cr)"}
        ],
        "auto_recipe": [ {"dur": 12, "target": 4}, {"dur": 12, "target": 4}, {"dur": 9, "target": 3}, {"dur": 6, "target": 2} ]
    },
    "balanced": {
        "name": "⚖️ Tridente",
        "desc": "13+13+13 días.",
        "blocks": [
            {"dur": 13, "cred": 5, "label": "Bloque 13d (5 Cr)"},
            {"dur": 13, "cred": 4, "label": "Bloque 13d (4 Cr)"}
        ],
        "auto_recipe": [ {"dur": 13, "target": 5}, {"dur": 13, "target": 4}, {"dur": 13, "target": 4} ]
    },
    "long": {
        "name": "✈️ Larga Estancia",
        "desc": "15+15+9 días.",
        "blocks": [
            {"dur": 15, "cred": 5, "label": "Gran Viaje 15d (5 Cr)"},
            {"dur": 9,  "cred": 3, "label": "Escapada 9d (3 Cr)"}
        ],
        "auto_recipe": [ {"dur": 15, "target": 5}, {"dur": 15, "target": 5}, {"dur": 9, "target": 3} ]
    },
    "micro": {
        "name": "🐜 Hormiga",
        "desc": "5x6 días + 1x9 días.",
        "blocks": [
            {"dur": 6, "cred": 2, "label": "Semana 6d (2 Cr)"},
            {"dur": 9, "cred": 3, "label": "Semana+ 9d (3 Cr)"}
        ],
        "auto_recipe": [ {"dur": 6, "target": 2}, {"dur": 6, "target": 2}, {"dur": 6, "target": 2}, {"dur": 6, "target": 2}, {"dur": 6, "target": 2}, {"dur": 9, "target": 3} ]
    },
    "sniper": {
        "name": "🎯 Francotirador",
        "desc": "13 días sueltos de guardia.",
        "blocks": [ {"dur": 1, "cred": 1, "label": "Día Suelto (1 Cr)"} ],
        "auto_recipe": [{"dur": 1, "target": 1}] * 13
    },
    "balanced_plus": {
        "name": "🧩 Flexible (4x8 + 1x7)",
        "desc": "4 de 8 días + 1 de 7 días.",
        "blocks": [
            {"dur": 8, "cred": 3, "label": "8d (3 Cr)"},
            {"dur": 8, "cred": 2, "label": "8d (2 Cr)"},
            {"dur": 7, "cred": 2, "label": "7d (2 Cr)"}
        ],
        "auto_recipe": [
            {"dur": 8, "target": 3}, {"dur": 8, "target": 3}, {"dur": 8, "target": 3},
            {"dur": 8, "target": 2}, {"dur": 7, "target": 2}
        ]
    }
}

DEFAULT_ROSTER = [
    {"ID_Puesto": "Jefe A",       "Nombre": "Jefe A",       "Turno": "A", "Rol": "Jefe",       "SV": False},
    {"ID_Puesto": "Subjefe A",    "Nombre": "Subjefe A",    "Turno": "A", "Rol": "Subjefe",    "SV": False},
    {"ID_Puesto": "Cond A",       "Nombre": "Cond A",       "Turno": "A", "Rol": "Conductor",  "SV": False},
    {"ID_Puesto": "Bombero A1",   "Nombre": "Bombero A1",   "Turno": "A", "Rol": "Bombero",    "SV": False},
    {"ID_Puesto": "Bombero A2",   "Nombre": "Bombero A2",   "Turno": "A", "Rol": "Bombero",    "SV": False},
    {"ID_Puesto": "Bombero A3",   "Nombre": "Bombero A3",   "Turno": "A", "Rol": "Bombero",    "SV": False},
    {"ID_Puesto": "Jefe B",       "Nombre": "Jefe B",       "Turno": "B", "Rol": "Jefe",       "SV": False},
    {"ID_Puesto": "Subjefe B",    "Nombre": "Subjefe B",    "Turno": "B", "Rol": "Subjefe",    "SV": False},
    {"ID_Puesto": "Cond B",       "Nombre": "Cond B",       "Turno": "B", "Rol": "Conductor",  "SV": False},
    {"ID_Puesto": "Bombero B1",   "Nombre": "Bombero B1",   "Turno": "B", "Rol": "Bombero",    "SV": False},
    {"ID_Puesto": "Bombero B2",   "Nombre": "Bombero B2",   "Turno": "B", "Rol": "Bombero",    "SV": False},
    {"ID_Puesto": "Bombero B3",   "Nombre": "Bombero B3",   "Turno": "B", "Rol": "Bombero",    "SV": False},
    {"ID_Puesto": "Jefe C",       "Nombre": "Jefe C",       "Turno": "C", "Rol": "Jefe",       "SV": False},
    {"ID_Puesto": "Subjefe C",    "Nombre": "Subjefe C",    "Turno": "C", "Rol": "Subjefe",    "SV": False},
    {"ID_Puesto": "Cond C",       "Nombre": "Cond C",       "Turno": "C", "Rol": "Conductor",  "SV": False},
    {"ID_Puesto": "Bombero C1",   "Nombre": "Bombero C1",   "Turno": "C", "Rol": "Bombero",    "SV": False},
    {"ID_Puesto": "Bombero C2",   "Nombre": "Bombero C2",   "Turno": "C", "Rol": "Bombero",    "SV": False},
    {"ID_Puesto": "Bombero C3",   "Nombre": "Bombero C3",   "Turno": "C", "Rol": "Bombero",    "SV": False},
]

# ==============================================================================
# 2. DEFINICIÓN DE TODAS LAS FUNCIONES (ANTES DE USARLAS)
# ==============================================================================

def load_data():
    if os.path.exists(DB_FILE):
        try:
            df = pd.read_csv(DB_FILE)
            if not df.empty:
                df['Inicio'] = pd.to_datetime(df['Inicio']).dt.date
                df['Fin'] = pd.to_datetime(df['Fin']).dt.date
        except: df = pd.DataFrame(columns=["Nombre", "Inicio", "Fin"])
    else:
        df = pd.DataFrame(columns=["Nombre", "Inicio", "Fin"])
    
    if os.path.exists(ADJ_FILE):
        try:
            adj_data = pd.read_csv(ADJ_FILE).to_dict('records')
        except: adj_data = []
    else:
        adj_data = []
    return df, adj_data

def save_data(df_vacs, list_adj):
    df_vacs.to_csv(DB_FILE, index=False)
    pd.DataFrame(list_adj).to_csv(ADJ_FILE, index=False)

def get_short_id(name, role, turn):
    if role == "Jefe": return f"J{turn}"
    if role == "Subjefe": return f"S{turn}"
    if role == "Conductor": return f"C{turn}"
    if "Bombero" in name:
        parts = name.split()
        if len(parts) > 1:
            suffix = parts[-1]
            if len(suffix) >= 2:
                return f"B{suffix[-1]}{turn}"
    return f"{name[:3]}{turn}"

def generate_night_template():
    wb = Workbook()
    ws = wb.active; ws.title = "Plan Nocturnas"
    ws.append(["Inicio (dd/mm/yyyy)", "Fin (dd/mm/yyyy)", "Notas"])
    ws.append(["2026-01-10", "2026-01-12", "Ejemplo"])
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out

def generate_base_schedule(year):
    is_leap = calendar.isleap(year)
    total_days = 366 if is_leap else 365
    status = {'A': 0, 'B': 2, 'C': 1} 
    schedule = {team: [] for team in TEAMS}
    for _ in range(total_days):
        for t in TEAMS:
            schedule[t].append('T' if status[t] == 0 else 'L')
            status[t] = (status[t] + 1) % 3
    return schedule, total_days

def is_in_night_period(day_idx, year, night_periods):
    current_date = datetime.date(year, 1, 1) + datetime.timedelta(days=day_idx)
    for start, end in night_periods:
        if start <= current_date <= end: return True
    return False

def get_night_transition_dates(night_periods):
    dates = set()
    for start, end in night_periods:
        dates.add(end) 
    return dates

def calculate_stats(roster_df, requests, year):
    base_sch, _ = generate_base_schedule(year)
    stats = {}
    for _, p in roster_df.iterrows():
        stats[p['Nombre']] = {'credits': 0, 'natural': 0}
    for req in requests:
        name = req['Nombre']
        if name not in stats: continue
        s_idx = req['Inicio'].timetuple().tm_yday - 1
        e_idx = req['Fin'].timetuple().tm_yday - 1
        row = roster_df[roster_df['Nombre'] == name].iloc[0]
        nat = (e_idx - s_idx) + 1
        cred = 0
        for d in range(s_idx, e_idx + 1):
            if base_sch[row['Turno']][d] == 'T': cred += 1
        stats[name]['credits'] += cred
        stats[name]['natural'] += nat
    return stats

def get_clustered_dates(available_idxs, needed_count):
    if not available_idxs: return []
    groups = []
    for k, g in groupby(enumerate(available_idxs), lambda ix: ix[0] - ix[1]):
        groups.append(list(map(itemgetter(1), g)))
    groups.sort(key=len, reverse=True)
    selected = []
    for group in groups:
        if len(selected) < needed_count:
            take = min(len(group), needed_count - len(selected))
            selected.extend(group[:take])
        else: break
    return sorted(selected)

def check_global_conflict_generic(start_idx, duration, person, occupation_map, base_sch, year, transition_dates):
    total_days = len(base_sch['A'])
    if start_idx + duration > total_days: return True
    for i in range(start_idx, start_idx + duration):
        d_obj = datetime.date(year, 1, 1) + timedelta(days=i)
        if d_obj in transition_dates:
            if base_sch[person['Turno']][i] == 'T': return True
        occupants = occupation_map.get(i, [])
        if len(occupants) >= 2: return True
        for occ in occupants:
            if occ['Turno'] == person['Turno']: return True
            if person['Rol'] != 'Bombero' and occ['Rol'] == person['Rol']: return True
    return False

def book_slot_gen(start_idx, duration, person, occupation_map):
    for i in range(start_idx, start_idx + duration):
        if i not in occupation_map: occupation_map[i] = []
        occupation_map[i].append(person)

def get_available_blocks_for_person(person_name, roster_df, current_requests, year, night_periods, month_range, strategy_key):
    base_sch, total_days = generate_base_schedule(year)
    transition_dates = get_night_transition_dates(night_periods)
    person = roster_df[roster_df['Nombre'] == person_name].iloc[0]
    start_month_idx = MESES.index(month_range[0]) + 1
    end_month_idx = MESES.index(month_range[1]) + 1
    occupation_map = {i:[] for i in range(total_days)}
    my_current_slots = [] 
    for req in current_requests:
        p_req = roster_df[roster_df['Nombre'] == req['Nombre']].iloc[0]
        s = req['Inicio'].timetuple().tm_yday - 1
        e = req['Fin'].timetuple().tm_yday - 1
        if req['Nombre'] != person_name:
            for d in range(s, e+1):
                if base_sch[p_req['Turno']][d] == 'T': occupation_map[d].append(p_req)
        else:
            my_current_slots.append((s, e))

    block_defs = STRATEGIES[strategy_key]['blocks']
    options = {b['label']: [] for b in block_defs}
    for d in range(total_days - 15): 
        d_date = datetime.date(year, 1, 1) + timedelta(days=d)
        if not (start_month_idx <= d_date.month <= end_month_idx): continue
        for b_def in block_defs:
            duration = b_def['dur']
            target_cred = b_def['cred']
            label_key = b_def['label']
            if not check_global_conflict_generic(d, duration, person, occupation_map, base_sch, year, transition_dates):
                overlap = False
                for ms in my_current_slots:
                    if not (d + duration - 1 < ms[0] - 2 or d > ms[1] + 2): overlap = True; break
                if not overlap:
                    credits = 0
                    for k in range(d, d+duration):
                        if base_sch[person['Turno']][k] == 'T': credits += 1
                    if credits == target_cred:
                        start_date = d_date
                        end_date = start_date + timedelta(days=duration-1)
                        txt = f"{start_date.strftime('%d/%m')} - {end_date.strftime('%d/%m')}"
                        options[label_key].append({'label': txt, 'start': start_date, 'end': end_date})
    return options

def auto_generate_schedule(roster_df, year, night_periods, strategy_key):
    base_sch, total_days = generate_base_schedule(year)
    transition_dates = get_night_transition_dates(night_periods)
    occupation_map = {} 
    generated_requests = []
    people = roster_df.to_dict('records')
    priority_order = ["Jefe", "Subjefe", "Conductor", "Bombero"]
    people.sort(key=lambda x: priority_order.index(x['Rol']))
    RECIPE = STRATEGIES[strategy_key]['auto_recipe']
    
    for person in people:
        my_slots = []
        current_recipe = RECIPE.copy()
        random.shuffle(current_recipe) 
        credits_got = 0
        
        for block in current_recipe:
            duration = block['dur']
            target = block['target']
            options = []
            for d in range(0, total_days - duration):
                c = sum([1 for k in range(d, d+duration) if base_sch[person['Turno']][k] == 'T'])
                if c == target:
                     if not check_global_conflict_generic(d, duration, person, occupation_map, base_sch, year, transition_dates):
                        options.append(d)
            random.shuffle(options)
            for start in options:
                overlap = any(start < s[0]+s[1]+2 and start+duration > s[0]-2 for s in my_slots)
                if not overlap:
                    book_slot_gen(start, duration, person, occupation_map)
                    my_slots.append((start, duration))
                    credits_got += target
                    generated_requests.append({
                        "Nombre": person['Nombre'],
                        "Inicio": datetime.date(year, 1, 1) + timedelta(days=start),
                        "Fin": datetime.date(year, 1, 1) + timedelta(days=start+duration-1)
                    })
                    break 
        
        if credits_got < 13:
            all_days_random = list(range(total_days))
            random.shuffle(all_days_random)
            for d in all_days_random:
                if credits_got >= 13: break
                if base_sch[person['Turno']][d] == 'T':
                    if not check_global_conflict_generic(d, 1, person, occupation_map, base_sch, year, transition_dates):
                        overlap = any(d < s[0]+s[1]+2 and d > s[0]-2 for s in my_slots)
                        if not overlap:
                            book_slot_gen(d, 1, person, occupation_map)
                            my_slots.append((d, 1))
                            credits_got += 1
                            generated_requests.append({
                                "Nombre": person['Nombre'],
                                "Inicio": datetime.date(year, 1, 1) + timedelta(days=d),
                                "Fin": datetime.date(year, 1, 1) + timedelta(days=d)
                            })
    return generated_requests

def render_annual_calendar(year, team, base_sch, night_periods, custom_schedule=None):
    html = f"<div style='font-family:monospace; font-size:10px;'>"
    html += """
    <div style='display:flex; gap:10px; margin-bottom:5px; font-size:11px; font-weight:bold;'>
        <span style='background:#d4edda; color:#155724; padding:2px 5px; border:1px solid #c3e6cb;'>T (Guardia)</span>
        <span style='background:#FFC000; color:#000; padding:2px 5px; border:1px solid #DAA520;'>V (Pedido)</span>
        <span style='background:#FFFFE0; color:#555; padding:2px 5px; border:1px solid #EEE8AA;'>V(R) (Relleno)</span>
        <span style='background:#28a745; color:white; padding:2px 5px;'>T (Noche)</span>
        <span style='border:2px solid red; padding:0px 5px; color:red;'>Fin Noche</span>
    </div>
    """
    html += "<div style='display:flex; margin-bottom:2px;'><div style='width:30px;'></div>"
    for d in range(1, 32):
        html += f"<div style='width:20px; text-align:center; color:#888;'>{d}</div>"
    html += "</div>"

    for m_idx, mes in enumerate(MESES):
        m_num = m_idx + 1
        days_in_month = calendar.monthrange(year, m_num)[1]
        html += f"<div style='display:flex; margin-bottom:2px;'><div style='width:30px; font-weight:bold;'>{mes}</div>"
        for d in range(1, 32):
            if d <= days_in_month:
                dt = datetime.date(year, m_num, d)
                d_idx = dt.timetuple().tm_yday - 1
                state = base_sch[team][d_idx]
                final_val = state
                if custom_schedule: final_val = custom_schedule[d_idx]

                bg_color = "#eee"; text_color = "#ccc"; border = "1px solid #fff"
                if final_val == 'T': 
                    bg_color = "#d4edda"; text_color = "#155724"
                    if is_in_night_period(d_idx, year, night_periods):
                        bg_color = "#28a745"; text_color = "white"
                elif final_val == 'V':
                    bg_color = "#FFC000"; text_color = "#000"
                elif final_val == 'V(R)':
                    bg_color = "#FFFFE0"; text_color = "#555"
                elif final_val == 'T+':
                    bg_color = "#ADD8E6"; text_color = "#000"
                elif final_val == 'L*':
                    bg_color = "#E6E6FA"; text_color = "#000"
                if dt in get_night_transition_dates(night_periods): border = "2px solid red"
                html += f"<div style='width:20px; background-color:{bg_color}; color:{text_color}; text-align:center; border:{border}; border-radius:2px;'>{state[0]}</div>"
            else:
                html += "<div style='width:20px;'></div>"
        html += "</div>"
    html += "</div>"
    return html

def get_candidates(person_missing, roster_df, day_idx, current_schedule, year, night_periods, adjustments_log_current_day=None):
    candidates = []
    missing_role = person_missing['Rol']
    missing_turn = person_missing['Turno']
    blocked_turns = set()
    if adjustments_log_current_day:
        for coverer_name in adjustments_log_current_day:
            cov_p = roster_df[roster_df['Nombre'] == coverer_name]
            if not cov_p.empty: blocked_turns.add(cov_p.iloc[0]['Turno'])
    turn_exhausted_from_night = None
    if day_idx > 0:
        prev_day_idx = day_idx - 1
        if is_in_night_period(prev_day_idx, year, night_periods):
            base_sch_temp, _ = generate_base_schedule(year)
            for t in TEAMS:
                if base_sch_temp[t][prev_day_idx] == 'T':
                    turn_exhausted_from_night = t; break
    for _, candidate in roster_df.iterrows():
        if candidate['Turno'] == missing_turn: continue
        cand_status = current_schedule[candidate['Nombre']][day_idx]
        if cand_status != 'L': continue 
        if candidate['Turno'] in blocked_turns: continue
        if turn_exhausted_from_night and candidate['Turno'] == turn_exhausted_from_night: continue
        is_compatible = False
        cand_role = candidate['Rol']
        if missing_role == "Jefe" and cand_role in ["Jefe", "Subjefe"]: is_compatible = True
        elif missing_role == "Subjefe" and cand_role in ["Jefe", "Subjefe"]: is_compatible = True
        elif missing_role == "Conductor" and (cand_role == "Conductor" or candidate['SV']): is_compatible = True
        elif missing_role == "Bombero" and (cand_role == "Bombero" or candidate['SV']): is_compatible = True
        if is_compatible: candidates.append(candidate['Nombre'])
    return candidates

def validate_and_generate_final(roster_df, requests, year, night_periods, forced_adjustments=None, strategy_key="standard"):
    if forced_adjustments is None: forced_adjustments = []
    base_schedule_turn, total_days = generate_base_schedule(year)
    final_schedule = {} 
    turn_coverage_counters = {'A': 0, 'B': 0, 'C': 0}
    person_coverage_counters = {name: 0 for name in roster_df['Nombre']}
    name_to_turn = {row['Nombre']: row['Turno'] for _, row in roster_df.iterrows()}
    
    for _, row in roster_df.iterrows():
        final_schedule[row['Nombre']] = base_schedule_turn[row['Turno']].copy()

    day_vacations = {i: [] for i in range(total_days)}
    natural_days_count = {name: 0 for name in roster_df['Nombre']}
    
    for req in requests:
        name = req['Nombre']
        s_idx = req['Inicio'].timetuple().tm_yday - 1
        e_idx = req['Fin'].timetuple().tm_yday - 1
        duration = (e_idx - s_idx) + 1
        natural_days_count[name] += duration
        for d in range(s_idx, e_idx + 1):
            if final_schedule[name][d] == 'T':
                day_vacations[d].append(name)
                final_schedule[name][d] = 'V'
            else:
                final_schedule[name][d] = 'V(L)'

    adjustments_log = []
    for d in range(total_days):
        absent_people = day_vacations[d]
        if not absent_people: continue
        current_day_coverers = []
        absent_people.sort(key=lambda x: 0 if "Jefe" in x or "Subjefe" in x else 1)

        for name_missing in absent_people:
            person_row = roster_df[roster_df['Nombre'] == name_missing].iloc[0]
            candidates = get_candidates(person_row, roster_df, d, final_schedule, year, night_periods, current_day_coverers)
            if candidates:
                valid = []
                for c in candidates:
                    prev = final_schedule[c][d-1] if d>0 else 'L'
                    prev2 = final_schedule[c][d-2] if d>1 else 'L'
                    if not (prev.startswith('T') and prev2.startswith('T')): valid.append(c)
                if valid:
                    valid.sort(key=lambda x: (turn_coverage_counters[name_to_turn[x]], person_coverage_counters[x], random.random()))
                    chosen = valid[0]
                    final_schedule[chosen][d] = f"T*({name_missing})"
                    adjustments_log.append((d, chosen, name_missing))
                    current_day_coverers.append(chosen)
                    turn_coverage_counters[name_to_turn[chosen]] += 1
                    person_coverage_counters[chosen] += 1

    for adj in forced_adjustments:
        d = adj['day_idx']
        p = adj['person']
        type_adj = adj['type']
        if type_adj == 'add': final_schedule[p][d] = "T+"
        elif type_adj == 'remove': final_schedule[p][d] = "L*"

    fill_log = {}
    for name in roster_df['Nombre']:
        if strategy_key == 'sniper':
            sched = final_schedule[name]
            for d in range(total_days - 2):
                if sched[d] == 'V':
                    if sched[d+1] == 'L': final_schedule[name][d+1] = 'V(R)'
                    if sched[d+2] == 'L': final_schedule[name][d+2] = 'V(R)'
        else:
            current = natural_days_count.get(name, 0)
            needed = 39 - current
            if needed > 0:
                available_idx = [i for i, x in enumerate(final_schedule[name]) if x == 'L']
                if len(available_idx) >= needed:
                    fill_idxs = get_clustered_dates(available_idx, needed)
                    for idx in fill_idxs:
                        final_schedule[name][idx] = 'V(R)'
    return final_schedule, adjustments_log, person_coverage_counters, fill_log

def get_work_days_count(final_schedule):
    counts = {}
    for name, sched in final_schedule.items():
        c = 0
        for s in sched:
            if s == 'T' or s.startswith('T*') or s == 'T+': c += 1
        counts[name] = c
    return counts

def find_adjustment_options(person_name, action_type, roster_df, year, night_periods, current_schedule):
    options = []
    base_sch, total_days = generate_base_schedule(year)
    transition_dates = get_night_transition_dates(night_periods)
    person_row = roster_df[roster_df['Nombre'] == person_name].iloc[0]
    vacation_counts = {i:0 for i in range(total_days)}
    for sched in current_schedule.values():
        for i, s in enumerate(sched):
            if 'V' in s: vacation_counts[i] += 1
    for d in range(total_days):
        current_status = current_schedule[person_name][d]
        if action_type == 'add':
            if current_status == 'L':
                if d > 0 and is_in_night_period(d-1, year, night_periods):
                    prev_t_turn = None
                    for t in TEAMS: 
                        if base_sch[t][d-1] == 'T': prev_t_turn = t
                    if prev_t_turn and person_row['Turno'] == prev_t_turn: continue
                if vacation_counts[d] < 2:
                    d_str = (datetime.date(year, 1, 1) + timedelta(days=d)).strftime("%d/%m")
                    options.append({'day_idx': d, 'label': f"{d_str} (Libre, {vacation_counts[d]} vacs)"})
        elif action_type == 'remove':
            if current_status == 'T' or current_status.startswith('T*'):
                if vacation_counts[d] == 0:
                    d_str = (datetime.date(year, 1, 1) + timedelta(days=d)).strftime("%d/%m")
                    tipo = "Guardia" if current_status == 'T' else "Cobertura"
                    options.append({'day_idx': d, 'label': f"{d_str} ({tipo} - Turno Completo)"})
    return options[:15]

def create_final_excel(schedule, roster_df, year, requests, fill_log, counters, night_periods, adjustments_log, strategy_key="standard"):
    wb = Workbook()
    s_T = PatternFill("solid", fgColor="C6EFCE"); s_V = PatternFill("solid", fgColor="FFC000")
    s_VR = PatternFill("solid", fgColor="FFFFE0"); s_Cov = PatternFill("solid", fgColor="FFC7CE")
    s_L = PatternFill("solid", fgColor="F2F2F2"); s_Night = PatternFill("solid", fgColor="A6A6A6")
    s_Extra = PatternFill("solid", fgColor="ADD8E6"); s_Free = PatternFill("solid", fgColor="E6E6FA")
    font_bold = Font(bold=True); font_red = Font(color="9C0006", bold=True)
    align_c = Alignment(horizontal="center", vertical="center")
    border_thin = Side(border_style="thin", color="000000")
    border_all = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    ws1 = wb.active; ws1.title = "Cuadrante"
    ws1.column_dimensions['A'].width = 20
    for i in range(2, 34): ws1.column_dimensions[get_column_letter(i)].width = 4
    
    curr_row = 1
    for t in TEAMS:
        ws1.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=32)
        cell_title = ws1.cell(curr_row, 1, f"TURNO {t}"); cell_title.font = Font(bold=True, color="FFFFFF"); cell_title.fill = PatternFill("solid", fgColor="000080"); cell_title.alignment = align_c
        curr_row += 2
        members = roster_df[roster_df['Turno'] == t].copy()
        role_order = ["Jefe", "Subjefe", "Conductor", "Bombero"]
        members['sort_key'] = members['Rol'].apply(lambda x: role_order.index(x))
        members = members.sort_values(by=['sort_key', 'Nombre'])
        for _, p in members.iterrows():
            nm = p['Nombre']; role = p['Rol']
            ws1.cell(curr_row, 1, f"{nm} ({role})").font = font_bold
            for d in range(1, 32): c = ws1.cell(curr_row, d+1, d); c.alignment = align_c; c.font = font_bold; c.border = border_all; c.fill = PatternFill("solid", fgColor="E0E0E0")
            curr_row += 1
            for m_idx, mes in enumerate(MESES):
                ws1.cell(curr_row, 1, mes).font = font_bold; ws1.cell(curr_row, 1).border = border_all
                d_month = calendar.monthrange(year, m_idx+1)[1]
                for d in range(1, 32):
                    cell = ws1.cell(curr_row, d+1); cell.border = border_all; cell.alignment = align_c
                    if d <= d_month:
                        dt = datetime.date(year, m_idx+1, d); d_y = dt.timetuple().tm_yday - 1
                        st_val = schedule[nm][d_y]
                        fill = s_L; val = ""
                        if st_val == 'T': fill = s_T; val = "T"
                        elif st_val == 'V': fill = s_V; val = "V"
                        elif st_val == 'V(R)': fill = s_VR; val = "v"
                        elif st_val == 'V(L)': fill = s_VR; val = "v"
                        elif st_val.startswith('T*'): 
                            fill = s_Cov; cell.font = font_red
                            raw_name = st_val.split('(')[1][:-1]
                            cov_p = roster_df[roster_df['Nombre'] == raw_name].iloc[0]
                            val = get_short_id(cov_p['Nombre'], cov_p['Rol'], cov_p['Turno'])
                        elif st_val == 'T+': fill = s_Extra; val = "T+"
                        elif st_val == 'L*': fill = s_Free; val = "L"
                        if is_in_night_period(d_y, year, night_periods): fill = s_Night
                        cell.fill = fill; cell.value = val
                    else: cell.fill = PatternFill("solid", fgColor="808080")
                curr_row += 1
            curr_row += 2 
    
    ws2 = wb.create_sheet("Estadísticas")
    headers = ["Nombre", "Turno", "Puesto", "Días Trabajados", "Gastado (T)", "Coberturas (T*)", "Total Vacs (Nat)"]
    ws2.append(headers)
    for _, p in roster_df.iterrows():
        name = p['Nombre']; sch = schedule[name]
        v_credits = sch.count('V')
        t_cover = counters[name]
        v_natural = sch.count('V') + sch.count('V(L)') + sch.count('V(R)')
        total_worked = 0
        for s in sch:
            if s == 'T' or s.startswith('T*') or s == 'T+': total_worked += 1
        ws2.append([name, p['Turno'], p['Rol'], total_worked, v_credits, t_cover, v_natural])

    ws4 = wb.create_sheet("Ajustes")
    ws4.append(["Fecha", "Cubre", "Ausente"])
    for d, c, a in adjustments_log:
        dt = datetime.date(year, 1, 1) + datetime.timedelta(days=d)
        ws4.append([dt.strftime("%d/%m/%Y"), c, a])
    
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out

# ==============================================================================
# MAIN EXECUTION
# ==============================================================================

st.title("🚒 Gestor V45.1: Cerebro Compartido")
st.markdown("**Diseñado por Marcos Esteban Vives**")

with st.expander("📘 MANUAL DE USUARIO (LÉEME)", expanded=True):
    st.markdown("""
    ### 🌐 CEREBRO COMPARTIDO
    Ahora la App guarda los datos en el servidor. **Lo que tú guardes, lo verán tus compañeros.**
    
    ### 0️⃣ REVISA LA PLANTILLA
    * Abre **"Plantilla"** (izquierda) y marca **SV** a los conductores.
    
    ### 1️⃣ CONFIGURACIÓN
    * **Nocturnas:** Descarga plantilla y sube Excel.
    
    ### 2️⃣ ASIGNA VACACIONES
    * **Modo Manual (Recomendado):** Elige una fecha y pulsa "Añadir". Se guardará al instante.
    * **Modo Automático:** Sobrescribirá TODO el cuadrante compartido. Úsalo con precaución.
    
    ### 3️⃣ EL NIVELADOR
    * Pulsa "🔄 Calcular Resultados" y ajusta los días. Los ajustes también se comparten.
    """)

current_requests_df, current_adjustments = load_data()
st.session_state.raw_requests_df = current_requests_df
st.session_state.forced_adjustments = current_adjustments
if 'locked_result' not in st.session_state: st.session_state.locked_result = None

current_requests = st.session_state.raw_requests_df.to_dict('records')
stats = calculate_stats(edited_df, current_requests, year_val)

# BARRA LATERAL
with st.sidebar:
    st.header("Configuración")
    year_val = st.number_input("Año", value=2026)
    
    with st.expander("Plantilla"):
        if 'roster_data' not in st.session_state:
            st.session_state.roster_data = pd.DataFrame(DEFAULT_ROSTER)
        column_cfg = {
            "ID_Puesto": st.column_config.TextColumn(disabled=True),
            "Turno": st.column_config.SelectboxColumn(options=TEAMS, required=True),
            "Rol": st.column_config.SelectboxColumn(options=ROLES, required=True),
            "SV": st.column_config.CheckboxColumn(label="¿Es SV?", help="Puede cubrir conductor", default=False)
        }
        edited_df = st.data_editor(
            st.session_state.roster_data, 
            column_config=column_cfg,
            use_container_width=True,
            key="roster_editor"
        )
        st.session_state.roster_data = edited_df
        
    with st.expander("Nocturnas"):
        if 'nights' not in st.session_state: st.session_state.nights = []
        c1, c2 = st.columns(2)
        dn_s = c1.date_input("Inicio", key="n_s", value=None)
        dn_e = c2.date_input("Fin", key="n_e", value=None)
        if st.button("Añadir Nocturna"):
            if dn_s and dn_e: st.session_state.nights.append((dn_s, dn_e))
        st.write(f"Periodos: {len(st.session_state.nights)}")
        
        st.download_button(label="⬇️ Descargar Plantilla Nocturnas", data=generate_night_template(), file_name="plantilla_nocturnas.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        uploaded_n = st.file_uploader("Excel Nocturnas", type=['xlsx'], key="n_up")
        if uploaded_n:
            try:
                df_n = pd.read_excel(uploaded_n)
                c = 0
                for _, row in df_n.iterrows():
                    try:
                        v1 = row.iloc[0]; v2 = row.iloc[1]
                        if not pd.isnull(v1) and not pd.isnull(v2):
                            d1 = pd.to_datetime(v1).date()
                            d2 = pd.to_datetime(v2).date()
                            st.session_state.nights.append((d1, d2)); c+=1
                    except: pass
                if c>0: st.success(f"Añadidos {c} periodos.")
            except: pass
        if st.button("Limpiar Nocturnas"): st.session_state.nights = []

    st.divider()
    if st.button("🔄 REFRESCAR DATOS", type="secondary"):
        st.rerun()

    def on_strategy_change():
        st.toast("⚠️ Estrategia visual cambiada.", icon="ℹ️")

    strategy_key = st.selectbox("🎯 Estrategia de Vacaciones", options=list(STRATEGIES.keys()), format_func=lambda x: STRATEGIES[x]['name'], on_change=on_strategy_change)
    st.info(STRATEGIES[strategy_key]['desc'])

    if st.button("🎲 Generar Automático (Sobrescribe)", type="primary"):
        with st.spinner("Generando..."):
            new_reqs = auto_generate_schedule(edited_df, year_val, st.session_state.nights, strategy_key)
            df_new = pd.DataFrame(new_reqs)
            save_data(df_new, []) # Reset ajustes
            st.session_state.raw_requests_df = df_new
            st.session_state.forced_adjustments = []
            st.session_state.locked_result = None 
        st.success("¡Hecho! Base de datos actualizada.")
        st.rerun()

# VISUAL
st.divider()
c_main, c_vis = st.columns([1, 2])

with c_main:
    st.subheader("2. Selección Manual")
    all_names = edited_df['Nombre'].tolist()
    names_sorted = sorted(all_names, key=lambda x: (0 if "Jefe" in x else 1 if "Subjefe" in x else 2 if "Cond" in x else 3))
    selected_person = st.selectbox("Selecciona Trabajador:", names_sorted)
    
    if selected_person:
        st.markdown("---")
        curr_stats = stats.get(selected_person, {'credits': 0, 'natural': 0})
        c = curr_stats['credits']
        remaining = 13 - c
        st.metric("Créditos Totales", f"{c} / 13", delta=remaining, delta_color="normal")
        
        recipe = STRATEGIES[strategy_key]['auto_recipe']
        req_counts = {}
        for item in recipe: req_counts[(item['dur'], item['target'])] = req_counts.get((item['dur'], item['target']), 0) + 1
        
        my_reqs = [r for r in current_requests if r['Nombre'] == selected_person]
        base_sch_temp, _ = generate_base_schedule(year_val)
        person_row = edited_df[edited_df['Nombre'] == selected_person].iloc[0]
        
        curr_counts = {}
        for r in my_reqs:
            dur = (r['Fin'] - r['Inicio']).days + 1
            s_idx = r['Inicio'].timetuple().tm_yday - 1
            cred_block = 0
            for d in range(s_idx, s_idx + dur):
                if base_sch_temp[person_row['Turno']][d] == 'T': cred_block += 1
            curr_counts[(dur, cred_block)] = curr_counts.get((dur, cred_block), 0) + 1
        
        sorted_keys = sorted(req_counts.keys(), key=lambda x: (-x[0], -x[1]))
        cols_puzzle = st.columns(len(sorted_keys))
        for idx, k in enumerate(sorted_keys):
            dur, cred = k
            total_needed = req_counts[k]
            have = curr_counts.get(k, 0)
            icon = "✅" if have >= total_needed else "⏳"
            with cols_puzzle[idx]:
                st.caption(f"{dur}d ({cred} Cr)")
                st.markdown(f"### {icon} {have}/{total_needed}")

        st.divider()
        if remaining <= 0:
            st.success("✅ Cupo cubierto.")
        else:
            month_range = st.select_slider("📅 Filtrar Meses:", options=MESES, value=(MESES[0], MESES[-1]))
            st.info(f"🔍 Buscando fichas disponibles...")
            options = get_available_blocks_for_person(selected_person, edited_df, current_requests, year_val, st.session_state.nights, month_range, strategy_key)
            block_defs = STRATEGIES[strategy_key]['blocks']
            tabs = st.tabs([b['label'] for b in block_defs])
            for i, b_def in enumerate(block_defs):
                key = b_def['label']
                with tabs[i]:
                    available_opts = options.get(key, [])
                    if not available_opts: st.warning("Sin opciones.")
                    else:
                        with st.container(height=200):
                            for opt in available_opts[:20]: 
                                if st.button(f"➕ {opt['label']}", key=f"add_{selected_person}_{opt['start']}_{i}"):
                                    df_now, adj_now = load_data()
                                    reqs_now = df_now.to_dict('records')
                                    reqs_now.append({"Nombre": selected_person, "Inicio": opt['start'], "Fin": opt['end']})
                                    save_data(pd.DataFrame(reqs_now), adj_now)
                                    st.session_state.raw_requests_df = pd.DataFrame(reqs_now)
                                    st.session_state.locked_result = None 
                                    st.rerun()

    st.markdown("---")
    st.write(f"**Mis Periodos:**")
    if not my_reqs: st.caption("Ninguno")
    else:
        for i, r in enumerate(my_reqs):
            c1, c2 = st.columns([4, 1])
            c1.write(f"{r['Inicio'].strftime('%d/%m')} - {r['Fin'].strftime('%d/%m')}")
            if c2.button("🗑️", key=f"del_{selected_person}_{i}"):
                df_now, adj_now = load_data()
                reqs_now = df_now.to_dict('records')
                for k, x in enumerate(reqs_now):
                    if x['Nombre'] == r['Nombre'] and x['Inicio'] == r['Inicio']:
                        del reqs_now[k]
                        break
                save_data(pd.DataFrame(reqs_now), adj_now)
                st.session_state.raw_requests_df = pd.DataFrame(reqs_now)
                st.session_state.locked_result = None
                st.rerun()

with c_vis:
    if selected_person:
        p_row = edited_df[edited_df['Nombre'] == selected_person].iloc[0]
        turn = p_row['Turno']
        st.subheader(f"3. Visor Turno {turn} ({selected_person})")
        
        base_sch, _ = generate_base_schedule(year_val)
        temp_sch = base_sch[turn].copy()
        my_reqs = [r for r in current_requests if r['Nombre'] == selected_person]
        for r in my_reqs:
            s = r['Inicio'].timetuple().tm_yday - 1
            e = r['Fin'].timetuple().tm_yday - 1
            for d in range(s, e+1):
                if temp_sch[d] == 'T': temp_sch[d] = 'V'
                else: temp_sch[d] = 'V(L)'
        
        if strategy_key == 'sniper':
             for d in range(len(temp_sch) - 2):
                 if temp_sch[d] == 'V' and temp_sch[d+1] == 'L': temp_sch[d+1] = 'V(R)'
                 if temp_sch[d] == 'V' and temp_sch[d+1] == 'V(R)' and temp_sch[d+2] == 'L': temp_sch[d+2] = 'V(R)'

        st.markdown(render_annual_calendar(year_val, turn, base_sch, st.session_state.nights, temp_sch), unsafe_allow_html=True)
    else:
        st.subheader("3. Visor Global")
        base_sch, _ = generate_base_schedule(year_val)
        st.markdown(render_annual_calendar(year_val, 'A', base_sch, st.session_state.nights), unsafe_allow_html=True)

st.divider()
st.header("⚙️ Ajuste Fino y Descarga")

if st.button("🔄 Calcular/Actualizar Resultados", type="primary"):
    with st.spinner("Calculando cuadrante final..."):
        df_now, adj_now = load_data()
        current_reqs = df_now.to_dict('records')
        sch, adj, count, fill = validate_and_generate_final(edited_df, current_reqs, year_val, st.session_state.nights, adj_now, strategy_key)
        excel_io = create_final_excel(sch, edited_df, year_val, current_reqs, fill, count, st.session_state.nights, adj, strategy_key)
        work_days = get_work_days_count(sch)
        st.session_state.locked_result = {
            "sch": sch, "adj": adj, "work_days": work_days, "excel": excel_io
        }
    st.success("¡Resultados actualizados!")

if st.session_state.locked_result:
    res = st.session_state.locked_result
    cols_eq = st.columns(3)
    for i, (name, count) in enumerate(res['work_days'].items()):
        with cols_eq[i % 3]:
            color = "green" if 121 <= count <= 123 else "red"
            st.markdown(f"**{name}**: <span style='color:{color}'>{count} días</span>", unsafe_allow_html=True)
    
    st.divider()
    col_poor, col_rich = st.columns(2)
    with col_poor:
        st.subheader("📉 Falta Jornada (<121)")
        poor_people = [n for n, c in res['work_days'].items() if c < 121]
        if not poor_people: st.success("Nadie necesita añadir.")
        else:
            p_select = st.selectbox("Seleccionar:", poor_people, key="sel_poor")
            if p_select:
                opts = find_adjustment_options(p_select, 'add', edited_df, year_val, st.session_state.nights, res['sch'])
                if opts:
                    day_opt = st.selectbox("Días Disponibles:", options=opts, format_func=lambda x: x['label'], key="opt_add")
                    if st.button(f"➕ Añadir a {p_select}"):
                        df_now, adj_now = load_data()
                        adj_now.append({'day_idx': day_opt['day_idx'], 'person': p_select, 'type': 'add'})
                        save_data(df_now, adj_now)
                        st.session_state.forced_adjustments = adj_now
                        st.session_state.locked_result = None 
                        st.rerun()
    with col_rich:
        st.subheader("📈 Sobra Jornada (>123)")
        rich_people = [n for n, c in res['work_days'].items() if c > 123]
        if not rich_people: st.success("Nadie necesita quitar.")
        else:
            r_select = st.selectbox("Seleccionar:", rich_people, key="sel_rich")
            if r_select:
                opts = find_adjustment_options(r_select, 'remove', edited_df, year_val, st.session_state.nights, res['sch'])
                if opts:
                    day_opt = st.selectbox("Días Disponibles:", options=opts, format_func=lambda x: x['label'], key="opt_rem")
                    if st.button(f"➖ Quitar a {r_select}"):
                        df_now, adj_now = load_data()
                        adj_now.append({'day_idx': day_opt['day_idx'], 'person': r_select, 'type': 'remove'})
                        save_data(df_now, adj_now)
                        st.session_state.forced_adjustments = adj_now
                        st.session_state.locked_result = None
                        st.rerun()

    st.divider()
    st.download_button("📥 Descargar Cuadrante Final", data=res['excel'], file_name=f"Cuadrante_Final_{year_val}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")
    st.markdown("**Diseñado por Marcos Esteban Vives**")
    st.caption("Asistente de programación. Esta información tiene un carácter meramente informativo. Para obtener asesoramiento o diagnóstico médicos, consulta a un profesional.")
else:
    st.info("Pulsa 'Calcular/Actualizar Resultados' para ver el estado de la plantilla y descargar.")
